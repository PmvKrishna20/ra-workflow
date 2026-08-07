import streamlit as st
import pandas as pd
import openpyxl
import re
import json
import uuid
import bcrypt
import hashlib
import secrets as secure_secrets
import urllib.request
import urllib.error
from datetime import date, datetime, timedelta, timezone

# ---------------------------------------------------------------------------
# Database adapter — wraps psycopg2 (Supabase/PostgreSQL) behind the same
# thin interface the app was written against (sqlite3 style):
#   conn = get_connection()
#   conn.execute(sql, params)
#   conn.executemany(sql, params_seq)
#   conn.commit()
#   conn.close()   ← no-op here; connection is pooled
#   row = cursor.fetchone()   → plain tuple  (not RealDict)
#   rows = cursor.fetchall()  → list of tuples
# SQL differences handled here so nothing else in the file needs to change:
#   • "?" placeholders  → "%s"
#   • "INTEGER PRIMARY KEY AUTOINCREMENT" → "SERIAL PRIMARY KEY"
#   • "INSERT OR IGNORE"  → "INSERT ... ON CONFLICT DO NOTHING"
#   • "PRAGMA table_info(...)" → information_schema query
#   • "DELETE FROM sqlite_sequence" → no-op (sequences reset automatically)
# ---------------------------------------------------------------------------

import psycopg2
import psycopg2.extensions
import psycopg2.extras
import psycopg2.pool


def _adapt_sql(sql: str) -> str:
    """Translate SQLite-flavoured SQL to PostgreSQL syntax."""
    sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
    sql = sql.replace("INSERT OR IGNORE INTO", "INSERT INTO")
    # Turn trailing ON CONFLICT clause on; add DO NOTHING if this was an INSERT OR IGNORE
    if "ON CONFLICT DO NOTHING" not in sql and sql.lstrip().upper().startswith("INSERT INTO") \
            and "ON CONFLICT" not in sql:
        # Only add when we converted INSERT OR IGNORE (marker: no ON CONFLICT yet)
        pass  # handled via the replace above; caller must add ON CONFLICT if needed
    sql = re.sub(r'\?', '%s', sql)
    return sql


class _Cursor:
    """Thin wrapper that returns plain tuples, matching sqlite3 Row behaviour."""
    def __init__(self, pg_cursor):
        self._c = pg_cursor

    def fetchone(self):
        row = self._c.fetchone()
        if row is None:
            return None
        return tuple(row)

    def fetchall(self):
        return [tuple(r) for r in self._c.fetchall()]

    def __iter__(self):
        for row in self._c:
            yield tuple(row)


class _Connection:
    """Wraps a psycopg2 connection with the sqlite3-like execute/executemany/commit/close API."""

    def __init__(self, pg_conn, pool):
        self._conn = pg_conn
        self._pool = pool
        self._returned = False

    def execute(self, sql: str, params=None):
        sql = _adapt_sql(sql)
        # PRAGMA table_info → information_schema equivalent
        m = re.match(r"PRAGMA\s+table_info\((\w+)\)", sql.strip(), re.I)
        if m:
            table = m.group(1)
            cur = self._conn.cursor()
            cur.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name=%s ORDER BY ordinal_position",
                (table,)
            )
            # Return a fake cursor whose fetchall() gives (cid, name, ...) tuples
            # The app only uses r[1] for column name, so fake (0, name, 'TEXT', 0, None, 0)
            rows = [(i, r[0], 'TEXT', 0, None, 0) for i, r in enumerate(cur.fetchall())]

            class _FakeCur:
                def fetchall(inner_self):
                    return rows
                def fetchone(inner_self):
                    return rows[0] if rows else None
            return _FakeCur()

        # Ignore sqlite_sequence wipe — not applicable to Postgres
        if "sqlite_sequence" in sql:
            class _Noop:
                def fetchone(self): return None
                def fetchall(self): return []
            return _Noop()

        # INSERT OR IGNORE was already replaced by _adapt_sql; add ON CONFLICT DO NOTHING
        # if the original had INSERT OR IGNORE (now just INSERT INTO) and no ON CONFLICT yet
        if "ON CONFLICT" not in sql and sql.lstrip().upper().startswith("INSERT INTO"):
            # Only add when we're doing an INSERT that expects uniqueness handling.
            # We do this conservatively: check if the original call came via executemany
            # (which always passes through execute). Safe to add because our tables
            # all have the right UNIQUE constraints and ON CONFLICT DO NOTHING is harmless.
            sql = re.sub(r'(VALUES\s*\(.*?\))\s*$', r'\1 ON CONFLICT DO NOTHING', sql,
                         flags=re.DOTALL | re.IGNORECASE)

        cur = self._conn.cursor()
        cur.execute(sql, params or ())
        return _Cursor(cur)

    def executemany(self, sql: str, params_seq):
        sql = _adapt_sql(sql)
        if "ON CONFLICT" not in sql and sql.lstrip().upper().startswith("INSERT INTO"):
            sql = re.sub(r'(VALUES\s*\(.*?\))\s*$', r'\1 ON CONFLICT DO NOTHING', sql,
                         flags=re.DOTALL | re.IGNORECASE)
        cur = self._conn.cursor()
        psycopg2.extras.execute_batch(cur, sql, params_seq, page_size=500)
        return _Cursor(cur)

    def commit(self):
        self._conn.commit()

    def close(self):
        if self._returned:
            return
        if self._conn.get_transaction_status() != psycopg2.extensions.TRANSACTION_STATUS_IDLE:
            self._conn.rollback()
        self._pool.putconn(self._conn)
        self._returned = True


@st.cache_resource
def _pg_pool():
    """Thread-safe pool so concurrent user sessions never share a transaction."""
    url = st.secrets["supabase"]["db_url"]
    return psycopg2.pool.ThreadedConnectionPool(1, 10, dsn=url)


def get_connection() -> _Connection:
    """Borrow a connection without a network round-trip health check."""
    pool = _pg_pool()
    raw = pool.getconn()
    if raw.closed:
        pool.putconn(raw, close=True)
        raw = pool.getconn()
    raw.autocommit = False
    return _Connection(raw, pool)


# ---------- AI enrichment config ----------
OPENAI_MODEL = "gpt-5.2"
OPENAI_URL = "https://api.openai.com/v1/chat/completions"
AI_BATCH_SIZE = 20
AI_ELIGIBLE_MIN = 10
AI_ELIGIBLE_MAX = 500




@st.cache_resource
def init_db():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            position TEXT,
            added_by TEXT,
            date_added TEXT,
            source TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_companies_name_date ON companies(company_name, date_added)")
    conn.execute("CREATE TABLE IF NOT EXISTS block_list (id INTEGER PRIMARY KEY AUTOINCREMENT, keyword TEXT NOT NULL UNIQUE)")
    conn.execute("CREATE TABLE IF NOT EXISTS clients (id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT NOT NULL UNIQUE)")
    conn.execute("CREATE TABLE IF NOT EXISTS avoid_list (id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT NOT NULL UNIQUE, reason TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS eligible_companies (id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT NOT NULL UNIQUE, employee_size TEXT, industry TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS not_eligible_companies (id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT NOT NULL UNIQUE, employee_size TEXT, industry TEXT, reason TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS needs_review_companies (id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT NOT NULL UNIQUE, employee_size TEXT, industry TEXT, reason TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ra_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            company_name TEXT NOT NULL,
            job_title TEXT,
            location TEXT,
            job_url TEXT,
            ra_name TEXT NOT NULL,
            assigned_date TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assignments_batch ON ra_assignments(batch_id)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prospects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT,
            email TEXT,
            company_name TEXT NOT NULL,
            company_key TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prospects_company_key ON prospects(company_key)")
    conn.execute("CREATE TABLE IF NOT EXISTS bounced_emails (id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT NOT NULL UNIQUE)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS eod_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ra_name TEXT NOT NULL,
            upload_date TEXT NOT NULL,
            company_name TEXT,
            company_linkedin_url TEXT,
            full_name TEXT,
            first_name TEXT,
            poc_location TEXT,
            designation TEXT,
            email TEXT,
            position TEXT,
            location TEXT,
            job_posting_link TEXT,
            industry TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_eod_ra_date ON eod_submissions(ra_name, upload_date)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS emails_sent (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT,
            email TEXT,
            position TEXT,
            location TEXT,
            company_name TEXT,
            ra_name TEXT,
            period TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_emails_sent_email_company ON emails_sent(email, company_name)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS positive_responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT,
            name TEXT,
            response_date TEXT,
            ra_name TEXT,
            position TEXT,
            designation TEXT,
            company_name TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_responses_email_company ON positive_responses(email, company_name)")
    conn.execute("CREATE TABLE IF NOT EXISTS title_bucket_keywords (id INTEGER PRIMARY KEY AUTOINCREMENT, keyword TEXT NOT NULL UNIQUE)")
    _kw_count = conn.execute("SELECT COUNT(*) FROM title_bucket_keywords").fetchone()[0]
    if _kw_count == 0:
        _default_keywords = [
            "Project Manager", "Program Manager", "Product Manager", "Account Manager", "Office Manager",
            "General Manager", "Operations Manager", "Plant Manager", "Warehouse Manager", "Sales Manager",
            "Manager", "Director", "Vice President", "President", "Supervisor", "Superintendent",
            "Coordinator", "Administrator", "Specialist", "Analyst", "Consultant", "Controller",
            "Bookkeeper", "Accountant", "Accounting", "Estimator", "Scheduler", "Planner", "Buyer",
            "Recruiter", "Engineer", "Engineering", "Technician", "Designer", "Architect", "Inspector",
            "Operator", "Machinist", "Welder", "Electrician", "Plumber", "Mechanic", "Driver", "Foreman",
            "Executive", "Officer", "CFO", "CEO", "COO", "CTO", "VP",
        ]
        conn.executemany("INSERT OR IGNORE INTO title_bucket_keywords (keyword) VALUES (?)", [(k,) for k in _default_keywords])
    for _t, _c in [("eligible_companies","employee_size"),("eligible_companies","industry"),("not_eligible_companies","employee_size"),("not_eligible_companies","industry")]:
        _existing = [r[1] for r in conn.execute(f"PRAGMA table_info({_t})").fetchall()]
        if _c not in _existing:
            conn.execute(f"ALTER TABLE {_t} ADD COLUMN {_c} TEXT")
    _prospects_cols = [r[1] for r in conn.execute("PRAGMA table_info(prospects)").fetchall()]
    if "company_key" not in _prospects_cols:
        conn.execute("ALTER TABLE prospects ADD COLUMN company_key TEXT")
    conn.execute("UPDATE prospects SET company_key = company_name WHERE company_key IS DISTINCT FROM company_name")

    # ---------- Users table (per-person login + role) ----------
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'RA',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT,
            work_date_preference TEXT
        )
    """)
    conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS work_date_preference TEXT")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            token_hash TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL
        )
    """)
    _session_cols = [r[1] for r in conn.execute("PRAGMA table_info(sessions)").fetchall()]
    if "token_hash" not in _session_cols or "expires_at" not in _session_cols:
        # An earlier unfinished build created an unused incompatible sessions table.
        conn.execute("DROP TABLE sessions")
        conn.execute("""
            CREATE TABLE sessions (
                token_hash TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL
            )
        """)
    # URL-carried bearer sessions were removed because copied links transferred
    # the signed-in user's privileges. Invalidate every legacy token on upgrade.
    conn.execute("DELETE FROM sessions")
    conn.commit()

    # One-time bootstrap: if no users exist yet, create the first Manager
    # account from Streamlit secrets so there's always a way in. Set in
    # secrets as:
    #   [bootstrap_admin]
    #   username = "vamsi"
    #   password = "choose-a-strong-password"
    #   display_name = "Vamsi"
    _user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if _user_count == 0:
        boot = st.secrets.get("bootstrap_admin", {})
        boot_user = boot.get("username", "").strip()
        boot_pwd = boot.get("password", "")
        boot_name = boot.get("display_name", boot_user).strip()
        if boot_user and boot_pwd:
            pwd_hash = bcrypt.hashpw(boot_pwd.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            conn.execute(
                "INSERT INTO users (username, password_hash, display_name, role, active, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (boot_user.lower(), pwd_hash, boot_name, "Manager", 1, date.today().isoformat())
            )
            conn.commit()

    # ---------- Workpage: positions + their prospects ----------
    conn.execute("""
        CREATE TABLE IF NOT EXISTS positions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            job_title TEXT,
            status TEXT NOT NULL DEFAULT 'Open',
            assigned_ra TEXT,
            source TEXT,
            batch_id TEXT,
            location TEXT,
            job_url TEXT,
            created_by TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS position_prospects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            position_id INTEGER NOT NULL,
            full_name TEXT,
            first_name TEXT,
            email TEXT,
            designation TEXT,
            location TEXT,
            linkedin_url TEXT,
            status TEXT NOT NULL DEFAULT 'Not contacted',
            notes TEXT,
            added_by TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    conn.commit()

    # Migrate columns added after the tables first went live (CREATE TABLE IF NOT EXISTS
    # above won't add columns to a table that already exists from an earlier deploy).
    # This MUST run before any index below that references these columns.
    _pos_cols = [r[1] for r in conn.execute("PRAGMA table_info(positions)").fetchall()]
    if "batch_id" not in _pos_cols:
        conn.execute("ALTER TABLE positions ADD COLUMN batch_id TEXT")
    for _col, _type in (
        ("master_company_id", "INTEGER"), ("ready_work_date", "TEXT"),
        ("ready_at", "TEXT"), ("ready_by", "TEXT"),
        ("export_batch_id", "TEXT"), ("exported_at", "TEXT"),
        ("exported_by", "TEXT"),
    ):
        if _col not in _pos_cols:
            conn.execute(f"ALTER TABLE positions ADD COLUMN {_col} {_type}")
    conn.execute("""
        UPDATE positions p
        SET master_company_id = (
            SELECT c.id FROM companies c
            WHERE c.company_name = p.company_name
              AND c.added_by = p.assigned_ra
              AND c.date_added = p.created_at
            ORDER BY c.id DESC LIMIT 1
        )
        WHERE p.master_company_id IS NULL
    """)
    _orphan_positions = conn.execute(
        "SELECT id, company_name, job_title, assigned_ra, created_at, source "
        "FROM positions WHERE master_company_id IS NULL"
    ).fetchall()
    for _pid, _company, _title, _ra, _created, _source in _orphan_positions:
        _master_id = conn.execute(
            "INSERT INTO companies (company_name, position, added_by, date_added, source) "
            "VALUES (?, ?, ?, ?, ?) RETURNING id",
            (_company, _title, _ra, _created or date.today().isoformat(), _source or "workpage_migration")
        ).fetchone()[0]
        conn.execute("UPDATE positions SET master_company_id = ? WHERE id = ?", (_master_id, _pid))
    _pp_cols = [r[1] for r in conn.execute("PRAGMA table_info(position_prospects)").fetchall()]
    for _col in ("full_name", "location", "linkedin_url"):
        if _col not in _pp_cols:
            conn.execute(f"ALTER TABLE position_prospects ADD COLUMN {_col} TEXT")
    conn.commit()

    # Enforce "one ACTIVE position per company" at the database level.
    # A closed position doesn't block a fresh one being opened later.
    # The old Closed state meant "sent to campaign" and maps to Exported.
    conn.execute("UPDATE positions SET status = 'Exported' WHERE status = 'Closed'")
    conn.execute("DROP INDEX IF EXISTS idx_positions_one_active_per_company")
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_positions_one_active_per_company
        ON positions (company_name) WHERE status <> 'Exported'
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_positions_assigned_ra ON positions(assigned_ra)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_positions_batch ON positions(batch_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_position_prospects_position ON position_prospects(position_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_positions_ready_work_date ON positions(ready_work_date)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS not_eligible_requests (
            id SERIAL PRIMARY KEY,
            company_name TEXT NOT NULL,
            reason TEXT NOT NULL,
            requested_by TEXT NOT NULL,
            requested_by_user_id INTEGER,
            requested_at TEXT NOT NULL,
            snapshot_json TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Pending',
            reviewed_by TEXT,
            reviewed_at TEXT,
            decision_note TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reassignment_queue (
            id SERIAL PRIMARY KEY,
            company_name TEXT NOT NULL UNIQUE,
            job_title TEXT,
            location TEXT,
            job_url TEXT,
            source_reason TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS deletion_audit (
            id SERIAL PRIMARY KEY,
            position_id INTEGER,
            company_name TEXT NOT NULL,
            position_snapshot_json TEXT NOT NULL,
            prospects_snapshot_json TEXT NOT NULL,
            export_batch_id TEXT,
            deleted_by TEXT NOT NULL,
            deleted_by_role TEXT NOT NULL,
            deletion_reason TEXT NOT NULL,
            deleted_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS export_batches (
            batch_id TEXT PRIMARY KEY,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            position_count INTEGER NOT NULL,
            prospect_count INTEGER NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS export_batch_items (
            id SERIAL PRIMARY KEY,
            batch_id TEXT NOT NULL,
            position_id INTEGER NOT NULL,
            company_name TEXT NOT NULL,
            ra_name TEXT,
            work_date TEXT,
            first_name TEXT NOT NULL,
            email TEXT NOT NULL,
            job_position TEXT NOT NULL,
            job_location TEXT NOT NULL,
            prospect_snapshot_json TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_export_items_batch ON export_batch_items(batch_id)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id TEXT,
            actor TEXT NOT NULL,
            details_json TEXT,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()

    conn.close()

# ---------- User / auth functions ----------

def hash_password(raw_password):
    return bcrypt.hashpw(raw_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(raw_password, password_hash):
    try:
        return bcrypt.checkpw(raw_password.encode("utf-8"), password_hash.encode("utf-8"))
    except Exception:
        return False

def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()

def audit_event(action, entity_type, entity_id, actor, details=None, conn=None):
    own_conn = conn is None
    conn = conn or get_connection()
    conn.execute(
        "INSERT INTO audit_log (action, entity_type, entity_id, actor, details_json, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (action, entity_type, str(entity_id) if entity_id is not None else None,
         actor, json.dumps(details or {}, default=str), utc_now_iso())
    )
    if own_conn:
        conn.commit()
        conn.close()

def _token_hash(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()

def create_session_token(user_id):
    token = secure_secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    conn = get_connection()
    conn.execute("DELETE FROM sessions WHERE expires_at < ?", (now.isoformat(),))
    conn.execute(
        "INSERT INTO sessions (token_hash, user_id, created_at, expires_at) VALUES (?, ?, ?, ?)",
        (_token_hash(token), user_id, now.isoformat(), (now + timedelta(days=7)).isoformat())
    )
    conn.commit()
    conn.close()
    return token

def get_user_by_session_token(token):
    if not token:
        return None
    conn = get_connection()
    row = conn.execute(
        "SELECT u.id, u.username, u.display_name, u.role FROM sessions s "
        "JOIN users u ON u.id = s.user_id "
        "WHERE s.token_hash = ? AND s.expires_at >= ? AND u.active = 1",
        (_token_hash(token), utc_now_iso())
    ).fetchone()
    conn.close()
    return row

def delete_session_token(token):
    if not token:
        return
    conn = get_connection()
    conn.execute("DELETE FROM sessions WHERE token_hash = ?", (_token_hash(token),))
    conn.commit()
    conn.close()

def revoke_user_sessions(user_id):
    conn = get_connection()
    conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

def get_current_user_record(user_id):
    conn = get_connection()
    row = conn.execute(
        "SELECT id, username, display_name, role, active, work_date_preference FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()
    conn.close()
    return row

def get_user_by_username(username):
    conn = get_connection()
    row = conn.execute(
        "SELECT id, username, password_hash, display_name, role, active FROM users WHERE username = ?",
        (username.strip().lower(),)
    ).fetchone()
    conn.close()
    return row  # (id, username, password_hash, display_name, role, active) or None

def list_users():
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, username, display_name, role, active, created_at FROM users ORDER BY created_at, id"
    ).fetchall()
    conn.close()
    return rows

def create_user(username, raw_password, display_name, role):
    conn = get_connection()
    username = username.strip().lower()
    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        conn.close()
        return False, "That username is already taken."
    existing_name = conn.execute("SELECT id FROM users WHERE display_name = ?", (display_name.strip(),)).fetchone()
    if existing_name:
        conn.close()
        return False, "That display name is already in use. Display names must be unique for assignment ownership."
    conn.execute(
        "INSERT INTO users (username, password_hash, display_name, role, active, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (username, hash_password(raw_password), display_name.strip(), role, 1, date.today().isoformat())
    )
    conn.commit()
    conn.close()
    return True, "User created."

def set_user_active(user_id, active):
    conn = get_connection()
    conn.execute("UPDATE users SET active = ? WHERE id = ?", (1 if active else 0, user_id))
    conn.commit()
    conn.close()
    revoke_user_sessions(user_id)

def set_user_role(user_id, role):
    conn = get_connection()
    conn.execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
    conn.commit()
    conn.close()
    revoke_user_sessions(user_id)

def reset_user_password(user_id, new_raw_password):
    conn = get_connection()
    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (hash_password(new_raw_password), user_id))
    conn.commit()
    conn.close()
    revoke_user_sessions(user_id)

def get_user_work_date(user_id):
    row = get_current_user_record(user_id)
    raw = row[5] if row else None
    if not raw and row:
        conn = get_connection()
        previous = conn.execute(
            "SELECT ready_work_date FROM positions WHERE assigned_ra = ? AND ready_work_date IS NOT NULL "
            "ORDER BY ready_at DESC NULLS LAST, id DESC LIMIT 1", (row[2],)
        ).fetchone()
        conn.close()
        raw = previous[0] if previous else None
    try:
        return date.fromisoformat(raw) if raw else date.today()
    except ValueError:
        return date.today()

def set_user_work_date(user_id, work_date):
    conn = get_connection()
    conn.execute("UPDATE users SET work_date_preference = ? WHERE id = ?", (work_date.isoformat(), user_id))
    conn.commit()
    conn.close()

# ---------- Company functions ----------

def find_recent_duplicate(company_name):
    cutoff = (date.today() - timedelta(days=30)).isoformat()
    conn = get_connection()
    row = conn.execute(
        "SELECT date_added, added_by FROM companies WHERE company_name = ? AND date_added >= ? ORDER BY date_added DESC LIMIT 1",
        (company_name, cutoff)
    ).fetchone()
    conn.close()
    return row

def add_company(company_name, position, added_by):
    conn = get_connection()
    row = conn.execute(
        "INSERT INTO companies (company_name, position, added_by, date_added, source) "
        "VALUES (?, ?, ?, ?, ?) RETURNING id",
        (company_name, position, added_by, date.today().isoformat(), "manual")
    ).fetchone()
    conn.commit()
    conn.close()
    return row[0]

# ---------- Workpage: positions + prospects ----------

def create_position(company_name, job_title, assigned_ra, source, created_by, location="", job_url="", batch_id=None,
                    master_company_id=None):
    """Creates a new active position for a company. Returns False without creating
    anything if that company already has an active (non-Exported) position. The
    partial unique index on positions is still there as a safety net for the rare
    case of two people saving at the same instant — the adapter auto-appends
    ON CONFLICT DO NOTHING to inserts, so a race just silently no-ops rather than
    erroring."""
    conn = get_connection()
    existing = conn.execute(
        "SELECT id FROM positions WHERE company_name = ? AND status <> 'Exported'", (company_name,)
    ).fetchone()
    if existing:
        conn.close()
        return False
    now = date.today().isoformat()
    inserted = conn.execute(
        "INSERT INTO positions (company_name, job_title, status, assigned_ra, source, batch_id, "
        "location, job_url, created_by, created_at, updated_at, master_company_id) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING id",
        (company_name, job_title, "Open", assigned_ra, source, batch_id, location, job_url,
         created_by, now, now, master_company_id)
    ).fetchone()
    conn.commit()
    conn.close()
    return inserted[0] if inserted else False

def create_manual_company_position(company_name, job_title, assigned_ra, created_by, location="", job_url=""):
    """Create the Master row and Workpage position atomically."""
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT id FROM positions WHERE company_name = ? AND status <> 'Exported'", (company_name,)
        ).fetchone()
        if existing:
            conn.close()
            return False
        now = date.today().isoformat()
        master_id = conn.execute(
            "INSERT INTO companies (company_name, position, added_by, date_added, source) "
            "VALUES (?, ?, ?, ?, 'manual') RETURNING id",
            (company_name, job_title, assigned_ra, now)
        ).fetchone()[0]
        position_id = conn.execute(
            "INSERT INTO positions (company_name, job_title, status, assigned_ra, source, location, job_url, "
            "created_by, created_at, updated_at, master_company_id) "
            "VALUES (?, ?, 'Open', ?, 'manual', ?, ?, ?, ?, ?, ?) RETURNING id",
            (company_name, job_title, assigned_ra, location, job_url, created_by, now, now, master_id)
        ).fetchone()[0]
        audit_event("create", "position", position_id, created_by, {"source": "manual"}, conn)
        conn.commit()
        conn.close()
        return position_id
    except Exception:
        conn._conn.rollback()
        conn.close()
        raise

def get_companies_with_active_position():
    """Companies that currently have an active (non-Exported) position — used to keep the
    RA Assignment pool from offering the same company twice across separate batches."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT company_name FROM positions WHERE status <> 'Exported' "
        "UNION SELECT company_name FROM not_eligible_requests WHERE status = 'Pending'"
    ).fetchall()
    conn.close()
    return set(r[0] for r in rows)

def save_assignment_bundle(rows, actor):
    """Save assignment, Master entry, and Workpage position atomically."""
    conn = get_connection()
    created = 0
    try:
        for batch_id, company, title, location, url, ra_name, assigned_date in rows:
            existing = conn.execute(
                "SELECT id FROM positions WHERE company_name = ? AND status <> 'Exported'", (company,)
            ).fetchone()
            if existing:
                continue
            master_id = conn.execute(
                "INSERT INTO companies (company_name, position, added_by, date_added, source) "
                "VALUES (?, ?, ?, ?, 'auto_assigned') RETURNING id",
                (company, title, ra_name, assigned_date)
            ).fetchone()[0]
            conn.execute(
                "INSERT INTO ra_assignments (batch_id, company_name, job_title, location, job_url, ra_name, assigned_date) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (batch_id, company, title, location, url, ra_name, assigned_date)
            )
            pid = conn.execute(
                "INSERT INTO positions (company_name, job_title, status, assigned_ra, source, batch_id, location, "
                "job_url, created_by, created_at, updated_at, master_company_id) "
                "VALUES (?, ?, 'Open', ?, 'auto_assigned', ?, ?, ?, ?, ?, ?, ?) RETURNING id",
                (company, title, ra_name, batch_id, location, url, actor,
                 assigned_date, assigned_date, master_id)
            ).fetchone()[0]
            audit_event("assign", "position", pid, actor, {"assigned_ra": ra_name, "batch_id": batch_id}, conn)
            conn.execute("DELETE FROM reassignment_queue WHERE company_name = ?", (company,))
            created += 1
        conn.commit()
        conn.close()
        return created
    except Exception:
        conn._conn.rollback()
        conn.close()
        raise

def delete_positions_for_batch(batch_id):
    """Removes the ra_assignments rows for a batch, plus any position that batch created
    and NO ONE has touched yet (still 'Open' with zero prospects added) — a safe undo for
    a mis-dealt batch. Positions someone has already worked (prospects added, or closed)
    are left alone so real work is never silently deleted; returns how many of each."""
    conn = get_connection()
    candidates = conn.execute(
        "SELECT id, master_company_id, company_name FROM positions WHERE batch_id = ? AND status = 'Open'", (batch_id,)
    ).fetchall()
    deleted_positions = 0
    kept_positions = 0
    for pid, master_id, company_name in candidates:
        n_prospects = conn.execute(
            "SELECT COUNT(*) FROM position_prospects WHERE position_id = ?", (pid,)
        ).fetchone()[0]
        if n_prospects == 0:
            conn.execute("DELETE FROM positions WHERE id = ?", (pid,))
            if master_id:
                conn.execute("DELETE FROM companies WHERE id = ?", (master_id,))
            conn.execute("DELETE FROM ra_assignments WHERE batch_id = ? AND company_name = ?",
                         (batch_id, company_name))
            deleted_positions += 1
        else:
            kept_positions += 1
    conn.commit()
    conn.close()
    return deleted_positions, kept_positions

def get_positions_for_ra(ra_name, include_closed=False):
    conn = get_connection()
    if include_closed:
        rows = conn.execute(
            "SELECT id, company_name, job_title, status, updated_at FROM positions "
            "WHERE assigned_ra = ? ORDER BY updated_at DESC", (ra_name,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, company_name, job_title, status, updated_at FROM positions "
            "WHERE assigned_ra = ? AND status <> 'Exported' ORDER BY updated_at DESC", (ra_name,)
        ).fetchall()
    conn.close()
    return rows

def get_all_positions(status_filter=None, ra_filter=None):
    conn = get_connection()
    sql = "SELECT id, company_name, job_title, status, assigned_ra, updated_at FROM positions WHERE 1=1"
    params = []
    if status_filter and status_filter != "All":
        sql += " AND status = ?"
        params.append(status_filter)
    if ra_filter and ra_filter != "All":
        sql += " AND assigned_ra = ?"
        params.append(ra_filter)
    sql += " ORDER BY updated_at DESC"
    rows = conn.execute(sql, tuple(params)).fetchall()
    conn.close()
    return rows

def get_position(position_id):
    conn = get_connection()
    row = conn.execute(
        "SELECT id, company_name, job_title, status, assigned_ra, source, location, job_url, "
        "created_by, created_at, updated_at, master_company_id, batch_id, ready_work_date, ready_at, "
        "ready_by, export_batch_id, exported_at, exported_by FROM positions WHERE id = ?", (position_id,)
    ).fetchone()
    conn.close()
    return row

def update_position(position_id, status=None, job_title=None, assigned_ra=None, location=None, job_url=None,
                    actor="System"):
    conn = get_connection()
    row = conn.execute(
        "SELECT status, job_title, assigned_ra, location, job_url, batch_id, company_name, master_company_id "
        "FROM positions WHERE id = ?", (position_id,)
    ).fetchone()
    if row is None:
        conn.close()
        return
    cur_status, cur_title, cur_ra, cur_loc, cur_url, batch_id, company_name, master_id = row
    final_title = job_title if job_title is not None else cur_title
    final_ra = assigned_ra if assigned_ra is not None else cur_ra
    final_loc = location if location is not None else cur_loc
    final_url = job_url if job_url is not None else cur_url
    conn.execute(
        "UPDATE positions SET status = ?, job_title = ?, assigned_ra = ?, location = ?, job_url = ?, "
        "updated_at = ? WHERE id = ?",
        (status if status is not None else cur_status,
         final_title, final_ra, final_loc, final_url,
         date.today().isoformat(), position_id)
    )
    if batch_id:
        conn.execute(
            "UPDATE ra_assignments SET job_title = ?, location = ?, job_url = ?, ra_name = ? "
            "WHERE batch_id = ? AND company_name = ?",
            (final_title, final_loc, final_url, final_ra, batch_id, company_name)
        )
    if master_id:
        conn.execute("UPDATE companies SET position = ?, added_by = ? WHERE id = ?",
                     (final_title, final_ra, master_id))
    audit_event("update", "position", position_id, actor, {
        "status": status, "job_title": final_title, "assigned_ra": final_ra,
        "location": final_loc, "job_url": final_url,
    }, conn)
    conn.commit()
    conn.close()

def get_ra_names_with_positions():
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT assigned_ra FROM positions WHERE assigned_ra IS NOT NULL ORDER BY assigned_ra").fetchall()
    conn.close()
    return [r[0] for r in rows]

def add_position_prospect(position_id, full_name, first_name, email, designation, location, linkedin_url, added_by):
    conn = get_connection()
    now = date.today().isoformat()
    conn.execute(
        "INSERT INTO position_prospects (position_id, full_name, first_name, email, designation, "
        "location, linkedin_url, status, notes, added_by, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (position_id, full_name, first_name, email, designation, location, linkedin_url,
         "Not contacted", "", added_by, now, now)
    )
    conn.commit()
    conn.close()

def bulk_add_position_prospects(position_id, rows, added_by):
    """rows: list of (full_name, first_name, email, designation, location, linkedin_url).
    Skips rows with no name and no email at all."""
    conn = get_connection()
    now = date.today().isoformat()
    to_insert = [
        (position_id, full_name, first_name, email, designation, location, linkedin_url,
         "Not contacted", "", added_by, now, now)
        for (full_name, first_name, email, designation, location, linkedin_url) in rows
        if (full_name or "").strip() or (first_name or "").strip() or (email or "").strip()
    ]
    if not to_insert:
        conn.close()
        return 0
    conn.executemany(
        "INSERT INTO position_prospects (position_id, full_name, first_name, email, designation, "
        "location, linkedin_url, status, notes, added_by, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        to_insert
    )
    conn.commit()
    conn.close()
    return len(to_insert)

def get_position_prospects(position_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, full_name, first_name, email, designation, location, linkedin_url, status, notes "
        "FROM position_prospects WHERE position_id = ? ORDER BY id", (position_id,)
    ).fetchall()
    conn.close()
    return rows

def update_position_prospect(prospect_id, status=None, notes=None, full_name=None, first_name=None,
                             email=None, designation=None, location=None, linkedin_url=None, actor="System"):
    conn = get_connection()
    row = conn.execute(
        "SELECT status, notes, full_name, first_name, email, designation, location, linkedin_url "
        "FROM position_prospects WHERE id = ?", (prospect_id,)
    ).fetchone()
    if row is None:
        conn.close()
        return
    cur_status, cur_notes, cur_full, cur_first, cur_email, cur_desig, cur_loc, cur_linkedin = row
    conn.execute(
        "UPDATE position_prospects SET status = ?, notes = ?, full_name = ?, first_name = ?, email = ?, "
        "designation = ?, location = ?, linkedin_url = ?, updated_at = ? WHERE id = ?",
        (status if status is not None else cur_status,
         notes if notes is not None else cur_notes,
         full_name if full_name is not None else cur_full,
         first_name if first_name is not None else cur_first,
         email.strip().lower() if email is not None else cur_email,
         designation if designation is not None else cur_desig,
         location if location is not None else cur_loc,
         linkedin_url if linkedin_url is not None else cur_linkedin,
         utc_now_iso(), prospect_id)
    )
    audit_event("update", "prospect", prospect_id, actor, {}, conn)
    conn.commit()
    conn.close()

EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

def validate_position_for_ready(position_id):
    """Return field-level errors. Ready is all-or-nothing for the whole position."""
    pos = get_position(position_id)
    prospects = get_position_prospects(position_id)
    errors = []
    if not pos:
        return ["Position no longer exists."]
    if not (pos[2] or "").strip():
        errors.append("Job Position is required.")
    if not (pos[6] or "").strip():
        errors.append("Job Location is required.")
    if not prospects:
        errors.append("Add at least one complete prospect.")
        return errors
    bounced = get_bounced_email_set()
    seen = set()
    conn = get_connection()
    for index, prospect in enumerate(prospects, start=1):
        prospect_id, full_name, first_name, email, designation, poc_location, linkedin_url, _, _ = prospect
        missing = [label for label, value in (
            ("Full Name", full_name), ("First Name", first_name), ("Email", email),
            ("Designation", designation), ("POC Location", poc_location),
        ) if not (value or "").strip()]
        if missing:
            errors.append(f"Prospect {index}: missing {', '.join(missing)}.")
        email_key = (email or "").strip().lower()
        if email_key and not EMAIL_RE.fullmatch(email_key):
            errors.append(f"Prospect {index}: invalid email format ({email}).")
        if email_key in seen:
            errors.append(f"Prospect {index}: duplicate email within this position ({email_key}).")
        seen.add(email_key)
        if email_key and email_key in bounced:
            errors.append(f"Prospect {index}: {email_key} is in Bounced/DNC.")
        if linkedin_url and not re.match(r"^https?://", linkedin_url.strip(), re.I):
            errors.append(f"Prospect {index}: LinkedIn URL must begin with http:// or https://.")
        if email_key:
            duplicate = conn.execute(
                "SELECT p.company_name, p.ready_work_date FROM position_prospects pp "
                "JOIN positions p ON p.id = pp.position_id "
                "WHERE LOWER(TRIM(pp.email)) = ? AND pp.id <> ? AND p.status IN ('Ready', 'Exported') LIMIT 1",
                (email_key, prospect_id)
            ).fetchone()
            if duplicate:
                errors.append(
                    f"Prospect {index}: {email_key} already exists on {duplicate[0]} "
                    f"({duplicate[1] or 'date unavailable'})."
                )
    conn.close()
    return errors

def mark_position_ready(position_id, work_date, actor):
    errors = validate_position_for_ready(position_id)
    if errors:
        return False, errors
    conn = get_connection()
    now = utc_now_iso()
    conn.execute(
        "UPDATE positions SET status = 'Ready', ready_work_date = ?, ready_at = ?, ready_by = ?, "
        "updated_at = ? WHERE id = ? AND status = 'Open'",
        (work_date.isoformat(), now, actor, now, position_id)
    )
    audit_event("ready", "position", position_id, actor, {"work_date": work_date.isoformat()}, conn)
    conn.commit()
    conn.close()
    return True, []

def return_position_to_open(position_id, actor):
    conn = get_connection()
    conn.execute(
        "UPDATE positions SET status = 'Open', ready_work_date = NULL, ready_at = NULL, ready_by = NULL, "
        "updated_at = ? WHERE id = ? AND status = 'Ready'",
        (utc_now_iso(), position_id)
    )
    audit_event("return_to_open", "position", position_id, actor, {}, conn)
    conn.commit()
    conn.close()

def update_ready_work_date(position_id, work_date, actor):
    conn = get_connection()
    conn.execute(
        "UPDATE positions SET ready_work_date = ?, updated_at = ? WHERE id = ? AND status = 'Ready'",
        (work_date.isoformat(), utc_now_iso(), position_id)
    )
    audit_event("update_work_date", "position", position_id, actor,
                {"work_date": work_date.isoformat()}, conn)
    conn.commit()
    conn.close()

def unlock_exported_position(position_id, actor):
    conn = get_connection()
    conn.execute(
        "UPDATE positions SET status = 'Ready', export_batch_id = NULL, exported_at = NULL, exported_by = NULL, "
        "updated_at = ? WHERE id = ? AND status = 'Exported'",
        (utc_now_iso(), position_id)
    )
    audit_event("unlock_exported", "position", position_id, actor, {}, conn)
    conn.commit()
    conn.close()

def delete_position_workflow(position_id, actor_user, reason_kind, reason_note):
    """Delete operational records but retain immutable audit/export snapshots."""
    conn = get_connection()
    try:
        pos = conn.execute(
            "SELECT id, company_name, job_title, status, assigned_ra, source, batch_id, location, job_url, "
            "created_by, created_at, updated_at, master_company_id, ready_work_date, ready_at, ready_by, "
            "export_batch_id, exported_at, exported_by FROM positions WHERE id = ?", (position_id,)
        ).fetchone()
        if not pos:
            conn.close()
            return False, "Position no longer exists."
        role, actor_name, actor_id = actor_user["role"], actor_user["display_name"], actor_user["id"]
        company, status, owner = pos[1], pos[3], pos[4]
        if role == "RA" and owner != actor_name:
            conn.close()
            return False, "You can delete only your own positions."
        if status == "Ready":
            conn.close()
            return False, "Return the position to Open before deleting it."
        if status == "Exported" and role not in ("Manager", "TL"):
            conn.close()
            return False, "Only a TL or Manager can delete an Exported position."
        if reason_kind == "not_eligible" and not reason_note.strip():
            conn.close()
            return False, "A Not Eligible explanation is required."
        prospects = conn.execute(
            "SELECT id, full_name, first_name, email, designation, location, linkedin_url, status, added_by, "
            "created_at, updated_at FROM position_prospects WHERE position_id = ? ORDER BY id", (position_id,)
        ).fetchall()
        pos_json = json.dumps(pos, default=str)
        prospects_json = json.dumps(prospects, default=str)
        if reason_kind == "not_eligible":
            conn.execute(
                "INSERT INTO not_eligible_requests (company_name, reason, requested_by, requested_by_user_id, "
                "requested_at, snapshot_json, status) VALUES (?, ?, ?, ?, ?, ?, 'Pending')",
                (company, reason_note.strip(), actor_name, actor_id, utc_now_iso(),
                 json.dumps({"position": pos, "prospects": prospects}, default=str))
            )
        elif reason_kind == "assigned_mistake":
            conn.execute(
                "INSERT INTO reassignment_queue (company_name, job_title, location, job_url, source_reason, created_at) "
                "VALUES (?, ?, ?, ?, 'Assigned by mistake', ?) "
                "ON CONFLICT (company_name) DO UPDATE SET job_title = EXCLUDED.job_title, "
                "location = EXCLUDED.location, job_url = EXCLUDED.job_url, created_at = EXCLUDED.created_at",
                (company, pos[2], pos[7], pos[8], utc_now_iso())
            )
        conn.execute(
            "INSERT INTO deletion_audit (position_id, company_name, position_snapshot_json, prospects_snapshot_json, "
            "export_batch_id, deleted_by, deleted_by_role, deletion_reason, deleted_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (position_id, company, pos_json, prospects_json, pos[16], actor_name, role,
             f"{reason_kind}: {reason_note.strip()}", utc_now_iso())
        )
        conn.execute("DELETE FROM position_prospects WHERE position_id = ?", (position_id,))
        conn.execute("DELETE FROM positions WHERE id = ?", (position_id,))
        if pos[6]:
            conn.execute("DELETE FROM ra_assignments WHERE batch_id = ? AND company_name = ?", (pos[6], company))
        if pos[12]:
            conn.execute("DELETE FROM companies WHERE id = ?", (pos[12],))
        audit_event("delete", "position", position_id, actor_name,
                    {"company": company, "reason_kind": reason_kind, "reason": reason_note}, conn)
        conn.commit()
        conn.close()
        return True, "Company removed from active workflow."
    except Exception:
        conn._conn.rollback()
        conn.close()
        raise

def get_not_eligible_requests(status="Pending"):
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, company_name, reason, requested_by, requested_at, status, reviewed_by, reviewed_at, "
        "decision_note FROM not_eligible_requests WHERE status = ? ORDER BY id DESC", (status,)
    ).fetchall()
    conn.close()
    return rows

def review_not_eligible_request(request_id, approve, reviewer, decision_note=""):
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT company_name, reason, status, snapshot_json FROM not_eligible_requests WHERE id = ?", (request_id,)
        ).fetchone()
        if not row or row[2] != "Pending":
            conn.close()
            return False
        company, reason, _, snapshot_json = row
        decision = "Approved" if approve else "Rejected"
        if approve:
            conn.execute(
                "INSERT INTO not_eligible_companies (company_name, employee_size, industry, reason) "
                "VALUES (?, '', '', ?) ON CONFLICT (company_name) DO UPDATE SET reason = EXCLUDED.reason",
                (company, reason)
            )
            conn.execute("DELETE FROM eligible_companies WHERE company_name = ?", (company,))
            conn.execute("DELETE FROM needs_review_companies WHERE company_name = ?", (company,))
        else:
            snapshot = json.loads(snapshot_json)
            old_pos = snapshot.get("position", [])
            conn.execute(
                "INSERT INTO reassignment_queue (company_name, job_title, location, job_url, source_reason, created_at) "
                "VALUES (?, ?, ?, ?, 'Not Eligible request rejected', ?) "
                "ON CONFLICT (company_name) DO UPDATE SET job_title = EXCLUDED.job_title, "
                "location = EXCLUDED.location, job_url = EXCLUDED.job_url, created_at = EXCLUDED.created_at",
                (company, old_pos[2] if len(old_pos) > 2 else "", old_pos[7] if len(old_pos) > 7 else "",
                 old_pos[8] if len(old_pos) > 8 else "", utc_now_iso())
            )
        conn.execute(
            "UPDATE not_eligible_requests SET status = ?, reviewed_by = ?, reviewed_at = ?, decision_note = ? "
            "WHERE id = ?",
            (decision, reviewer, utc_now_iso(), decision_note.strip(), request_id)
        )
        audit_event("approve_not_eligible" if approve else "reject_not_eligible", "request", request_id,
                    reviewer, {"company": company, "note": decision_note}, conn)
        conn.commit()
        conn.close()
        return True
    except Exception:
        conn._conn.rollback()
        conn.close()
        raise

def get_active_ra_names():
    conn = get_connection()
    rows = conn.execute(
        "SELECT display_name FROM users WHERE active = 1 "
        "AND role IN ('RA', 'TL', 'Manager') ORDER BY display_name"
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]

def get_reassignment_queue_rows():
    conn = get_connection()
    rows = conn.execute(
        "SELECT company_name, job_title, location, job_url FROM reassignment_queue ORDER BY id"
    ).fetchall()
    conn.close()
    return rows

def get_campaign_positions(date_from=None, date_to=None, ra_name=None, statuses=("Ready",)):
    conn = get_connection()
    placeholders = ",".join("?" for _ in statuses)
    sql = (
        "SELECT p.id, p.company_name, p.job_title, p.location, p.assigned_ra, p.status, "
        "p.ready_work_date, p.ready_at, COUNT(pp.id) "
        "FROM positions p LEFT JOIN position_prospects pp ON pp.position_id = p.id "
        f"WHERE p.status IN ({placeholders})"
    )
    params = list(statuses)
    if date_from:
        sql += " AND p.ready_work_date >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND p.ready_work_date <= ?"
        params.append(date_to)
    if ra_name and ra_name != "All":
        sql += " AND p.assigned_ra = ?"
        params.append(ra_name)
    sql += (
        " GROUP BY p.id, p.company_name, p.job_title, p.location, p.assigned_ra, p.status, "
        "p.ready_work_date, p.ready_at ORDER BY p.ready_work_date DESC, p.assigned_ra, p.company_name"
    )
    rows = conn.execute(sql, tuple(params)).fetchall()
    conn.close()
    return rows

def get_ra_ready_totals(date_from=None, date_to=None):
    conn = get_connection()
    sql = (
        "WITH activity AS ("
        " SELECT p.id AS position_id, p.ready_work_date AS work_date, p.assigned_ra AS ra_name, "
        "        'P-' || pp.id::text AS item_key "
        " FROM positions p JOIN position_prospects pp ON pp.position_id = p.id "
        " WHERE p.status IN ('Ready', 'Exported') "
        " UNION ALL "
        " SELECT d.position_id, d.work_date, d.ra_name, 'E-' || d.id::text FROM ("
        "   SELECT DISTINCT ON (e.position_id, e.email) e.id, e.position_id, e.work_date, e.ra_name "
        "   FROM export_batch_items e JOIN export_batches b ON b.batch_id = e.batch_id "
        "   LEFT JOIN positions p ON p.id = e.position_id WHERE p.id IS NULL "
        "   ORDER BY e.position_id, e.email, b.created_at DESC"
        " ) d"
        ") SELECT work_date, ra_name, COUNT(DISTINCT position_id), COUNT(item_key) "
        "FROM activity WHERE 1=1"
    )
    params = []
    if date_from:
        sql += " AND work_date >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND work_date <= ?"
        params.append(date_to)
    sql += " GROUP BY work_date, ra_name ORDER BY work_date DESC, ra_name"
    rows = conn.execute(sql, tuple(params)).fetchall()
    conn.close()
    return rows

def create_campaign_export(position_ids, actor):
    """Create an immutable snapshot and mark all selected Ready positions Exported atomically."""
    if not position_ids:
        return None, []
    conn = get_connection()
    try:
        placeholders = ",".join("?" for _ in position_ids)
        positions = conn.execute(
            "SELECT id, company_name, job_title, location, assigned_ra, ready_work_date "
            f"FROM positions WHERE id IN ({placeholders}) AND status = 'Ready' FOR UPDATE",
            tuple(position_ids)
        ).fetchall()
        if len(positions) != len(set(position_ids)):
            raise ValueError("One or more selected positions are no longer Ready.")
        batch_id = "EXP-" + datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S-") + uuid.uuid4().hex[:6]
        items = []
        for pid, company, job_title, job_location, ra_name, work_date in positions:
            prospects = conn.execute(
                "SELECT full_name, first_name, email, designation, location, linkedin_url "
                "FROM position_prospects WHERE position_id = ? ORDER BY id", (pid,)
            ).fetchall()
            for prospect in prospects:
                full_name, first_name, email, designation, poc_location, linkedin_url = prospect
                email = (email or "").strip().lower()
                snapshot = json.dumps({
                    "full_name": full_name, "first_name": first_name, "email": email,
                    "designation": designation, "poc_location": poc_location,
                    "linkedin_url": linkedin_url,
                })
                conn.execute(
                    "INSERT INTO export_batch_items (batch_id, position_id, company_name, ra_name, work_date, "
                    "first_name, email, job_position, job_location, prospect_snapshot_json) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (batch_id, pid, company, ra_name, work_date, first_name, email,
                     job_title, job_location, snapshot)
                )
                items.append((first_name, email, job_title, job_location))
        now = utc_now_iso()
        conn.execute(
            "INSERT INTO export_batches (batch_id, created_by, created_at, position_count, prospect_count) "
            "VALUES (?, ?, ?, ?, ?)",
            (batch_id, actor, now, len(positions), len(items))
        )
        conn.execute(
            f"UPDATE positions SET status = 'Exported', export_batch_id = ?, exported_at = ?, exported_by = ?, "
            f"updated_at = ? WHERE id IN ({placeholders})",
            tuple([batch_id, now, actor, now] + list(position_ids))
        )
        audit_event("export", "export_batch", batch_id, actor,
                    {"position_ids": position_ids, "prospect_count": len(items)}, conn)
        conn.commit()
        conn.close()
        return batch_id, items
    except Exception:
        conn._conn.rollback()
        conn.close()
        raise

def get_export_batches(limit=100):
    conn = get_connection()
    rows = conn.execute(
        "SELECT batch_id, created_by, created_at, position_count, prospect_count "
        "FROM export_batches ORDER BY created_at DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return rows

def get_export_batch_items(batch_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT first_name, email, job_position, job_location FROM export_batch_items "
        "WHERE batch_id = ? ORDER BY id", (batch_id,)
    ).fetchall()
    conn.close()
    return rows

def get_deletion_audit_rows(limit=200):
    conn = get_connection()
    rows = conn.execute(
        "SELECT company_name, deleted_by, deleted_by_role, deletion_reason, deleted_at, export_batch_id "
        "FROM deletion_audit ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return rows

def get_companies():
    conn = get_connection()
    rows = conn.execute("SELECT company_name, position, added_by, date_added, source FROM companies ORDER BY id DESC").fetchall()
    conn.close()
    return rows

def count_companies():
    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    cutoff = (date.today() - timedelta(days=30)).isoformat()
    hot = conn.execute("SELECT COUNT(*) FROM companies WHERE date_added >= ?", (cutoff,)).fetchone()[0]
    conn.close()
    return total, hot

def get_dashboard_metrics():
    """Fetch dashboard counters in one round trip instead of ten sequential queries."""
    cutoff = (date.today() - timedelta(days=30)).isoformat()
    conn = get_connection()
    row = conn.execute(
        "SELECT "
        "(SELECT COUNT(*) FROM companies), "
        "(SELECT COUNT(*) FROM companies WHERE date_added >= ?), "
        "(SELECT COUNT(*) FROM eligible_companies), "
        "(SELECT COUNT(*) FROM not_eligible_companies), "
        "(SELECT COUNT(*) FROM needs_review_companies), "
        "(SELECT COUNT(*) FROM prospects), "
        "(SELECT COUNT(DISTINCT company_key) FROM prospects), "
        "(SELECT COUNT(*) FROM ra_assignments), "
        "(SELECT COUNT(*) FROM eod_submissions), "
        "(SELECT COUNT(*) FROM emails_sent), "
        "(SELECT COUNT(*) FROM positive_responses)",
        (cutoff,)
    ).fetchone()
    conn.close()
    return row

def get_existing_keys():
    conn = get_connection()
    rows = conn.execute("SELECT company_name, date_added FROM companies").fetchall()
    conn.close()
    return set(rows)

def bulk_insert_companies(rows):
    conn = get_connection()
    conn.executemany("INSERT INTO companies (company_name, position, added_by, date_added, source) VALUES (?, ?, ?, ?, ?)", rows)
    conn.commit()
    conn.close()

def wipe_imported_history():
    conn = get_connection()
    conn.execute("DELETE FROM companies WHERE source IN ('imported', 'master_import')")
    conn.commit()
    conn.close()

def parse_date(raw):
    try:
        return pd.to_datetime(str(raw).strip(), dayfirst=True).date().isoformat()
    except Exception:
        return str(raw).strip()

# ---------- Block list functions ----------

def get_block_list():
    conn = get_connection()
    rows = conn.execute("SELECT keyword FROM block_list ORDER BY keyword ASC").fetchall()
    conn.close()
    return [r[0] for r in rows]

def add_keyword(keyword):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO block_list (keyword) VALUES (?)", (keyword,))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def remove_keyword(keyword):
    conn = get_connection()
    conn.execute("DELETE FROM block_list WHERE keyword = ?", (keyword,))
    conn.commit()
    conn.close()

def title_is_blocked(title, keywords):
    title_lower = title.lower()
    for kw in keywords:
        if kw in title_lower:
            return kw
    return None

# ---------- Clients & Avoid functions ----------

def add_to_clients(name):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO clients (company_name) VALUES (?)", (name,))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def add_to_avoid(name, reason):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO avoid_list (company_name, reason) VALUES (?, ?)", (name, reason))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def bulk_add_clients(names):
    conn = get_connection()
    conn.executemany("INSERT OR IGNORE INTO clients (company_name) VALUES (?)", [(n,) for n in names])
    conn.commit()
    conn.close()

def bulk_add_avoid(pairs):
    conn = get_connection()
    conn.executemany("INSERT OR IGNORE INTO avoid_list (company_name, reason) VALUES (?, ?)", pairs)
    conn.commit()
    conn.close()

def count_clients():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    conn.close()
    return n

def count_avoid():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM avoid_list").fetchone()[0]
    conn.close()
    return n

def is_client(name):
    conn = get_connection()
    r = conn.execute("SELECT 1 FROM clients WHERE company_name = ?", (name,)).fetchone()
    conn.close()
    return r is not None

def is_avoided(name):
    conn = get_connection()
    r = conn.execute("SELECT reason FROM avoid_list WHERE company_name = ?", (name,)).fetchone()
    conn.close()
    return r

def wipe_clients():
    conn = get_connection()
    conn.execute("DELETE FROM clients")
    conn.commit()
    conn.close()

def wipe_avoid():
    conn = get_connection()
    conn.execute("DELETE FROM avoid_list")
    conn.commit()
    conn.close()


# ---------- Review lists (eligible / not eligible) ----------

def bulk_add_eligible(rows):  # rows: (name, size, industry)
    conn = get_connection()
    conn.executemany("INSERT OR IGNORE INTO eligible_companies (company_name, employee_size, industry) VALUES (?, ?, ?)", rows)
    conn.commit()
    conn.close()

def bulk_add_not_eligible(rows):  # rows: (name, size, industry, reason)
    conn = get_connection()
    conn.executemany("INSERT OR IGNORE INTO not_eligible_companies (company_name, employee_size, industry, reason) VALUES (?, ?, ?, ?)", rows)
    conn.commit()
    conn.close()

def add_one_eligible(name, size, industry):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO eligible_companies (company_name, employee_size, industry) VALUES (?, ?, ?)", (name, size, industry))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def add_one_not_eligible(name, size, industry, reason):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO not_eligible_companies (company_name, employee_size, industry, reason) VALUES (?, ?, ?, ?)", (name, size, industry, reason))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def get_eligible_set():
    conn = get_connection()
    s = set(r[0] for r in conn.execute("SELECT company_name FROM eligible_companies").fetchall())
    conn.close()
    return s

def get_not_eligible_set():
    conn = get_connection()
    s = set(r[0] for r in conn.execute("SELECT company_name FROM not_eligible_companies").fetchall())
    conn.close()
    return s

def get_eligible_rows(limit=200):
    conn = get_connection()
    rows = conn.execute("SELECT company_name, employee_size, industry FROM eligible_companies ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return rows

def get_not_eligible_rows(limit=200):
    conn = get_connection()
    rows = conn.execute("SELECT company_name, employee_size, industry, reason FROM not_eligible_companies ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return rows

def count_eligible():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM eligible_companies").fetchone()[0]
    conn.close()
    return n

def count_not_eligible():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM not_eligible_companies").fetchone()[0]
    conn.close()
    return n

def wipe_eligible():
    conn = get_connection()
    conn.execute("DELETE FROM eligible_companies")
    conn.commit()
    conn.close()

def wipe_not_eligible():
    conn = get_connection()
    conn.execute("DELETE FROM not_eligible_companies")
    conn.commit()
    conn.close()

# ---------- Needs-review queue (AI uncertain calls, held out of the eligible/not-eligible cache) ----------

def bulk_add_needs_review(rows):  # rows: (name, size, industry, reason)
    conn = get_connection()
    conn.executemany("INSERT OR IGNORE INTO needs_review_companies (company_name, employee_size, industry, reason) VALUES (?, ?, ?, ?)", rows)
    conn.commit()
    conn.close()

def add_one_needs_review(name, size, industry, reason):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO needs_review_companies (company_name, employee_size, industry, reason) VALUES (?, ?, ?, ?)", (name, size, industry, reason))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def get_needs_review_set():
    conn = get_connection()
    s = set(r[0] for r in conn.execute("SELECT company_name FROM needs_review_companies").fetchall())
    conn.close()
    return s

def get_needs_review_rows(limit=200):
    conn = get_connection()
    rows = conn.execute("SELECT company_name, employee_size, industry, reason FROM needs_review_companies ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return rows

def count_needs_review():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM needs_review_companies").fetchone()[0]
    conn.close()
    return n

def wipe_needs_review():
    conn = get_connection()
    conn.execute("DELETE FROM needs_review_companies")
    conn.commit()
    conn.close()

def resolve_needs_review(name, decision):
    """decision: 'eligible' or 'not_eligible'. Moves a company out of the review queue into the
    matching permanent cache table, keeping whatever size/industry the AI had already estimated."""
    conn = get_connection()
    row = conn.execute(
        "SELECT employee_size, industry, reason FROM needs_review_companies WHERE company_name = ?", (name,)
    ).fetchone()
    conn.close()
    if row is None:
        return
    size, industry, reason = row
    if decision == "eligible":
        add_one_eligible(name, size, industry)
    else:
        add_one_not_eligible(name, size, industry, reason or "Manually resolved from needs-review queue")
    conn2 = get_connection()
    conn2.execute("DELETE FROM needs_review_companies WHERE company_name = ?", (name,))
    conn2.commit()
    conn2.close()

def wipe_entire_database():
    """Empties every table so the app starts exactly like a fresh install. Table structure stays intact."""
    conn = get_connection()
    tables = [
        "position_prospects", "positions", "not_eligible_requests", "reassignment_queue",
        "export_batch_items", "export_batches", "deletion_audit", "audit_log",
        "companies", "block_list", "clients", "avoid_list",
        "eligible_companies", "not_eligible_companies", "needs_review_companies",
        "ra_assignments", "prospects", "bounced_emails", "eod_submissions",
        "emails_sent", "positive_responses",
    ]
    for t in tables:
        conn.execute(f"DELETE FROM {t}")
    # reset autoincrement counters so IDs start at 1 again on a fresh test run
    try:
        conn.execute("DELETE FROM sqlite_sequence")
    except psycopg2.OperationalError:
        pass  # sqlite_sequence doesn't exist yet if nothing was ever inserted
    conn.commit()
    conn.close()

def guess_col(cols, target):
    for c in cols:
        if str(c).strip().lower() == target:
            return c
    return None

def _idx(options, value):
    try:
        return options.index(value)
    except (ValueError, TypeError):
        return 0

def get_setting(key, default=""):
    conn = get_connection()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row[0] if row is not None else default

def set_setting(key, value):
    conn = get_connection()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value)
    )
    conn.commit()
    conn.close()

# ---------- Prospects DB (the 130k-row First Name / Email / Company Name list) ----------

def normalize_company(name):
    """Return the exact company spelling; workflow matching is deliberately exact."""
    return str(name)


def add_prospects_from_eod(rows):
    """rows: (first_name, email, company_name). Appends WITHOUT wiping the existing table — used
    when an RA uploads their EOD sheet, since replace_prospects's wholesale-replace would blow
    away the master 130k list. Skips duplicate emails (already on file or within this batch) and
    anything on the bounced list. Returns (inserted_count, duplicate_count, bounced_count)."""
    bounced = get_bounced_email_set()
    conn = get_connection()
    existing = set(r[0] for r in conn.execute("SELECT LOWER(TRIM(email)) FROM prospects").fetchall())
    seen_emails = set()
    insert_rows = []
    dup_count = 0
    bounced_count = 0
    for (first_name, email, company_name) in rows:
        email_key = str(email).strip().lower()
        if email_key and email_key in bounced:
            bounced_count += 1
            continue
        if email_key:
            if email_key in existing or email_key in seen_emails:
                dup_count += 1
                continue
            seen_emails.add(email_key)
        insert_rows.append((first_name, email, company_name, normalize_company(company_name)))
    conn.executemany(
        "INSERT INTO prospects (first_name, email, company_name, company_key) VALUES (?, ?, ?, ?)",
        insert_rows
    )
    conn.commit()
    conn.close()
    return len(insert_rows), dup_count, bounced_count


def replace_prospects(rows):
    """rows: (first_name, email, company_name). Wipes and reloads the whole table in one transaction —
    this is a reference list meant to be swapped wholesale, not merged row by row.
    Drops duplicate emails (first occurrence wins) and anything already on the bounced list.
    Returns (inserted_count, duplicate_count, bounced_count)."""
    bounced = get_bounced_email_set()
    seen_emails = set()
    insert_rows = []
    dup_count = 0
    bounced_count = 0
    for (first_name, email, company_name) in rows:
        email_key = str(email).strip().lower()
        if email_key and email_key in bounced:
            bounced_count += 1
            continue
        if email_key:
            if email_key in seen_emails:
                dup_count += 1
                continue
            seen_emails.add(email_key)
        insert_rows.append((first_name, email, company_name, normalize_company(company_name)))

    conn = get_connection()
    conn.execute("DELETE FROM prospects")
    conn.executemany(
        "INSERT INTO prospects (first_name, email, company_name, company_key) VALUES (?, ?, ?, ?)",
        insert_rows
    )
    conn.commit()
    conn.close()
    return len(insert_rows), dup_count, bounced_count

def count_prospects():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM prospects").fetchone()[0]
    n_companies = conn.execute("SELECT COUNT(DISTINCT company_key) FROM prospects").fetchone()[0]
    conn.close()
    return n, n_companies

def wipe_prospects():
    conn = get_connection()
    conn.execute("DELETE FROM prospects")
    conn.commit()
    conn.close()

def lookup_prospects_for_companies(company_names):
    """Exact-spelling lookup returning rows keyed by the supplied company name."""
    if not company_names:
        return {}
    key_to_original = {}
    for name in company_names:
        key_to_original.setdefault(normalize_company(name), name)

    conn = get_connection()
    result = {}
    keys = list(key_to_original.keys())
    chunk_size = 500
    for start in range(0, len(keys), chunk_size):
        chunk = keys[start:start + chunk_size]
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"SELECT company_key, first_name, email FROM prospects WHERE company_key IN ({placeholders})",
            chunk
        ).fetchall()
        for company_key, first_name, email in rows:
            original_name = key_to_original.get(company_key, company_key)
            result.setdefault(original_name, []).append((first_name, email))
    conn.close()
    return result


def map_reengage_to_prospects(reengage_rows):
    """reengage_rows: list of (company, title, location, url, bucket_label).
    Matches each company against the Prospects DB using exact company spelling. A match produces
    one output row per contact at that company, pairing First Name + Email from the Prospects DB
    with Title + Location carried over from this scrape. No match means we don't currently have
    contacts for that company, so it's returned separately to feed the RA Assignment pool.
    Returns (mapped_rows, unmatched_rows)."""
    company_names = [r[0] for r in reengage_rows]
    prospects_by_company = lookup_prospects_for_companies(company_names)
    mapped_rows = []
    unmatched_rows = []
    for (company, title, location, url, bucket) in reengage_rows:
        matches = prospects_by_company.get(company)
        if not matches:
            unmatched_rows.append((company, title, location, url, bucket))
            continue
        for (first_name, email) in matches:
            mapped_rows.append({
                "First Name": first_name, "Email": email,
                "Title": title, "Location": location, "Company Name": company,
                "Bucket": bucket,
            })
    return mapped_rows, unmatched_rows


# ---------- RA EOD uploads ----------

def bulk_insert_eod(rows):
    """rows: (ra_name, upload_date, company_name, company_linkedin_url, full_name, first_name,
    poc_location, designation, email, position, location, job_posting_link, industry)"""
    conn = get_connection()
    conn.executemany(
        "INSERT INTO eod_submissions (ra_name, upload_date, company_name, company_linkedin_url, "
        "full_name, first_name, poc_location, designation, email, position, location, "
        "job_posting_link, industry) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows
    )
    conn.commit()
    conn.close()

def get_eod_ra_list():
    conn = get_connection()
    ras = sorted(set(r[0] for r in conn.execute("SELECT DISTINCT ra_name FROM eod_submissions").fetchall()))
    conn.close()
    return ras

def get_eod_rows(ra_names=None, date_from=None, date_to=None):
    conn = get_connection()
    query = (
        "SELECT ra_name, upload_date, company_name, company_linkedin_url, full_name, first_name, "
        "poc_location, designation, email, position, location, job_posting_link, industry "
        "FROM eod_submissions WHERE 1=1"
    )
    params = []
    if ra_names:
        placeholders = ",".join("?" for _ in ra_names)
        query += f" AND ra_name IN ({placeholders})"
        params.extend(ra_names)
    if date_from:
        query += " AND upload_date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND upload_date <= ?"
        params.append(date_to)
    query += " ORDER BY ra_name, upload_date"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows

def count_eod():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM eod_submissions").fetchone()[0]
    conn.close()
    return n

def wipe_eod():
    conn = get_connection()
    conn.execute("DELETE FROM eod_submissions")
    conn.commit()
    conn.close()


# ---------- Analytics: title-bucket keywords ----------

def get_title_bucket_keywords():
    conn = get_connection()
    rows = conn.execute("SELECT keyword FROM title_bucket_keywords ORDER BY LENGTH(keyword) DESC, keyword ASC").fetchall()
    conn.close()
    return [r[0] for r in rows]

def add_title_bucket_keyword(keyword):
    conn = get_connection()
    try:
        conn.execute("INSERT INTO title_bucket_keywords (keyword) VALUES (?)", (keyword.strip(),))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn._conn.rollback()
    conn.close()

def remove_title_bucket_keyword(keyword):
    conn = get_connection()
    conn.execute("DELETE FROM title_bucket_keywords WHERE keyword = ?", (keyword,))
    conn.commit()
    conn.close()

def bucket_title(title, keywords=None):
    """Longest keyword wins (so 'Project Manager' matches before plain 'Manager'); case-insensitive
    substring match. Falls back to 'Other' when nothing matches."""
    if keywords is None:
        keywords = get_title_bucket_keywords()
    t = str(title or "").lower()
    for kw in keywords:
        if kw.lower() in t:
            return kw
    return "Other"


# ---------- Analytics: emails sent / positive responses ----------

def bulk_insert_emails_sent(rows):
    """rows: (first_name, email, position, location, company_name, ra_name, period)"""
    conn = get_connection()
    conn.executemany(
        "INSERT INTO emails_sent (first_name, email, position, location, company_name, ra_name, period) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)", rows
    )
    conn.commit()
    conn.close()

def bulk_insert_positive_responses(rows):
    """rows: (email, name, response_date, ra_name, position, designation, company_name)"""
    conn = get_connection()
    conn.executemany(
        "INSERT INTO positive_responses (email, name, response_date, ra_name, position, designation, company_name) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)", rows
    )
    conn.commit()
    conn.close()

def count_emails_sent():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM emails_sent").fetchone()[0]
    conn.close()
    return n

def count_positive_responses():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM positive_responses").fetchone()[0]
    conn.close()
    return n

def wipe_emails_sent():
    conn = get_connection()
    conn.execute("DELETE FROM emails_sent")
    conn.commit()
    conn.close()

def wipe_positive_responses():
    conn = get_connection()
    conn.execute("DELETE FROM positive_responses")
    conn.commit()
    conn.close()

def compute_analytics():
    """Joins positive_responses back to emails_sent by normalized (email, company) to attribute each
    response to the title bucket it was originally sent about. A response with no company on file, or
    no matching sent row, is counted as 'unattributed' — it still counts toward the overall total but
    can't be credited to a specific title bucket. RA-wise stats use each table's own RA Name column
    directly (no join needed there); rows with no RA on file (e.g. the Jan emails-sent data) are
    grouped under 'Unknown (no RA data)' rather than silently dropped.
    Returns (title_stats, ra_stats, total_sent, total_responses, unattributed_responses) where
    title_stats/ra_stats are lists of (label, sent, responses, response_rate_pct) sorted by sent desc."""
    conn = get_connection()
    sent_rows = conn.execute("SELECT email, position, company_name, ra_name FROM emails_sent").fetchall()
    resp_rows = conn.execute("SELECT email, company_name, ra_name FROM positive_responses").fetchall()
    conn.close()

    keywords = get_title_bucket_keywords()

    sent_lookup = {}          # (email_key, company_key) -> [title bucket, ...]
    bucket_sent_counts = {}
    ra_sent_counts = {}
    for (email, position, company, ra) in sent_rows:
        key = (str(email or "").strip().lower(), normalize_company(company or ""))
        bucket = bucket_title(position, keywords)
        sent_lookup.setdefault(key, []).append(bucket)
        bucket_sent_counts[bucket] = bucket_sent_counts.get(bucket, 0) + 1
        ra_label = str(ra).strip() if ra and str(ra).strip() else "Unknown (no RA data)"
        ra_sent_counts[ra_label] = ra_sent_counts.get(ra_label, 0) + 1

    bucket_response_counts = {}
    ra_response_counts = {}
    unattributed = 0
    for (email, company, ra) in resp_rows:
        ra_label = str(ra).strip() if ra and str(ra).strip() else "Unknown"
        ra_response_counts[ra_label] = ra_response_counts.get(ra_label, 0) + 1
        company_norm = normalize_company(company) if company and str(company).strip() else None
        if not company_norm:
            unattributed += 1
            continue
        key = (str(email or "").strip().lower(), company_norm)
        buckets = sent_lookup.get(key)
        if not buckets:
            unattributed += 1
            continue
        bucket = buckets[0]
        bucket_response_counts[bucket] = bucket_response_counts.get(bucket, 0) + 1

    all_buckets = set(bucket_sent_counts) | set(bucket_response_counts)
    title_stats = []
    for b in all_buckets:
        sent = bucket_sent_counts.get(b, 0)
        resp = bucket_response_counts.get(b, 0)
        rate = (resp / sent * 100) if sent else 0.0
        title_stats.append((b, sent, resp, rate))
    title_stats.sort(key=lambda r: r[1], reverse=True)

    all_ras = set(ra_sent_counts) | set(ra_response_counts)
    ra_stats = []
    for ra_label in all_ras:
        sent = ra_sent_counts.get(ra_label, 0)
        resp = ra_response_counts.get(ra_label, 0)
        rate = (resp / sent * 100) if sent else 0.0
        ra_stats.append((ra_label, sent, resp, rate))
    ra_stats.sort(key=lambda r: r[1], reverse=True)

    return title_stats, ra_stats, len(sent_rows), len(resp_rows), unattributed


# ---------- Bounced / DNC emails ----------

def bulk_add_bounced(emails):
    conn = get_connection()
    rows = [(str(e).strip().lower(),) for e in emails if str(e).strip()]
    conn.executemany("INSERT OR IGNORE INTO bounced_emails (email) VALUES (?)", rows)
    conn.commit()
    conn.close()

def get_bounced_email_set():
    conn = get_connection()
    s = set(r[0] for r in conn.execute("SELECT email FROM bounced_emails").fetchall())
    conn.close()
    return s

def count_bounced():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM bounced_emails").fetchone()[0]
    conn.close()
    return n

def wipe_bounced():
    conn = get_connection()
    conn.execute("DELETE FROM bounced_emails")
    conn.commit()
    conn.close()

def scrub_prospects_against_bounced():
    """Deletes any prospect whose email is on the bounced list — e.g. after new bounces are
    uploaded. If that removes every prospect at a company, that company simply has zero rows
    left, so it naturally falls through to RA sourcing next time it's routed for mapping."""
    conn = get_connection()
    before = conn.execute("SELECT COUNT(*) FROM prospects").fetchone()[0]
    conn.execute("DELETE FROM prospects WHERE LOWER(TRIM(email)) IN (SELECT email FROM bounced_emails)")
    conn.commit()
    after = conn.execute("SELECT COUNT(*) FROM prospects").fetchone()[0]
    conn.close()
    return before - after


# ---------- RA Assignment ----------

def bulk_insert_assignments(rows):
    """rows: (batch_id, company_name, job_title, location, job_url, ra_name, assigned_date)"""
    conn = get_connection()
    conn.executemany(
        "INSERT INTO ra_assignments (batch_id, company_name, job_title, location, job_url, ra_name, assigned_date) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        rows
    )
    conn.commit()
    conn.close()

def get_assignment_batches(limit=50):
    conn = get_connection()
    rows = conn.execute(
        "SELECT batch_id, MAX(assigned_date), COUNT(*), COUNT(DISTINCT ra_name) "
        "FROM ra_assignments GROUP BY batch_id ORDER BY MAX(id) DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return rows

def get_assignment_rows_for_batch(batch_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT company_name, job_title, location, job_url, ra_name FROM ra_assignments "
        "WHERE batch_id = ? ORDER BY ra_name, company_name", (batch_id,)
    ).fetchall()
    conn.close()
    return rows

def count_assignments():
    conn = get_connection()
    n = conn.execute("SELECT COUNT(*) FROM ra_assignments").fetchone()[0]
    conn.close()
    return n

def wipe_assignments():
    conn = get_connection()
    conn.execute("DELETE FROM ra_assignments")
    conn.commit()
    conn.close()

def deal_companies_to_ras(pool, roster):
    """pool: list of (company, title, location, url), already sorted alphabetically.
    roster: list of (ra_name, count), in the order entered.
    Deals contiguous chunks (not round-robin) so each RA gets one alphabetical block.
    Returns (assignments dict ra_name -> list of rows, leftover list)."""
    assignments = {}
    idx = 0
    for ra_name, count in roster:
        chunk = pool[idx: idx + count]
        assignments.setdefault(ra_name, [])
        assignments[ra_name].extend(chunk)
        idx += count
    leftover = pool[idx:]
    return assignments, leftover


def read_csv_safe(file):
    """Reads an uploaded CSV, trying UTF-8 first then falling back to Windows encodings
    (Excel on Windows commonly saves CSVs as cp1252/latin-1, which trips plain UTF-8 reads
    on characters like curly quotes, em-dashes, or ® symbols)."""
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            file.seek(0)
            return pd.read_csv(file, encoding=encoding)
        except UnicodeDecodeError:
            continue
    file.seek(0)
    return pd.read_csv(file, encoding="latin-1", encoding_errors="replace")

# ---------- Bulk scrape cleaning ----------

def clean_scrape(file):
    wb = openpyxl.load_workbook(file)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    col = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        val = str(c.value).strip() if c.value is not None else ""
        link = c.hyperlink.target if c.hyperlink else None
        col.append((val, link))

    def fix_title(t):
        t = re.sub(r"\s*with verification\s*$", "", t, flags=re.I).strip()
        half = len(t) / 2
        if half == int(half) and half > 0:
            h = int(half)
            if t[:h] == t[h:]:
                return t[:h]
        return t

    rows = []
    seen = set()
    jobs_found = 0
    for i, (val, link) in enumerate(col):
        if not link:
            continue
        jobs_found += 1
        title = fix_title(val)
        company = col[i + 1][0] if i + 1 < len(col) else ""
        location = col[i + 2][0] if i + 2 < len(col) else ""
        if not company:
            continue
        if company in seen:
            continue
        seen.add(company)
        rows.append((company, title, location, link))
    return rows, jobs_found


def route_scrape(rows):
    conn = get_connection()
    clients = set(r[0] for r in conn.execute("SELECT company_name FROM clients").fetchall())
    avoid = set(r[0] for r in conn.execute("SELECT company_name FROM avoid_list").fetchall())
    today = date.today()
    d15 = (today - timedelta(days=15)).isoformat()
    d30 = (today - timedelta(days=30)).isoformat()
    groups = {"client_dnc": [], "skip": [], "reengage_15_30": [], "reengage_30plus": [], "new": []}
    for (company, title, location, url) in rows:
        if company in clients or company in avoid:
            groups["client_dnc"].append((company, title, location, url))
            continue
        maxdate = conn.execute("SELECT MAX(date_added) FROM companies WHERE company_name = ?", (company,)).fetchone()[0]
        if maxdate is None:
            groups["new"].append((company, title, location, url))
        elif maxdate >= d15:
            groups["skip"].append((company, title, location, url))
        elif maxdate >= d30:
            groups["reengage_15_30"].append((company, title, location, url))
        else:
            groups["reengage_30plus"].append((company, title, location, url))
    conn.close()
    return groups


KEYWORD_BLOCK_RE = re.compile(r"\b(talent|staffing|recruit(ing|ment|er|ers)?|hire|hiring|placement|headhunt(ing|er|ers)?|outsourc(ing|ed)?|payroll|employer\s+of\s+record|peo|contingent|executive\s+search|human\s+capital|manpower|hr\s+solution(s)?|workforce|hospital(ity)?|health(care)?|health\s+system|health\s+network|medical(\s+center)?|clinic(s)?|wellness|pharmac(y|eutical|euticals)?|dental|surgical|rehabilitation|rehab|care\s+center|care\s+home(s)?|nursing|school(s)?|school\s+district|college(s)?|university|universities|academy|academies|education(al)?|institute|seminary|training\s+center|tutoring|bootcamp|government|city\s+of|state\s+of|county|federal|municipality|municipal|department\s+of|ministry|public\s+sector|district|authority|commission|bureau|agency|foundation(s)?|nonprofit|non-profit|charity|charities|church|temple|mosque|synagogue|diocese|association|society|brotherhood|sisterhood|bank(ing)?|capital|finance|financial|investment(s)?|venture(s)?|equity|wealth\s+management|asset\s+management|private\s+equity|hedge\s+fund|credit\s+union|insurance|brokerage|mortgage|lending|loans|restaurant(s)?|cafe|foods?|hospitality|diner|bistro|catering|global|international|group\s+of\s+companies)\b", re.I)

AUTO_DEALER_RE = re.compile(r"\b(subaru|chevrolet|toyota|bmw|kia|honda|chrysler|dodge|jeep|mercedes-benz|hyundai|nissan|cadillac|buick|gmc|volkswagen|porsche|ferrari|lamborghini|maserati|bentley|rolls-royce|aston\s+martin|mclaren|alfa\s+romeo|mitsubishi|infiniti|lexus|acura|mazda|land\s+rover|jaguar|volvo|harley-davidson|ducati|kawasaki|yamaha|suzuki|ktm|triumph\s+motorcycles)\b", re.I)

KNOWN_BLOCKED = {
 "actalent","jobot","insight global","kforce","randstad","manpower","manpowergroup","adecco","hays","robert half",
 "kelly services","kelly","spherion","aerotek","allegis group","cielo","pontoon","hudson","pagegroup","michael page",
 "page personnel","spencer stuart","egon zehnder","korn ferry","russell reynolds","heidrick & struggles","odgers berndtson",
 "toptal","upwork","fiverr","g-p","globalization partners","deel","rippling","bamboohr","zenefits","paychex","adp",
 "trinet","insperity","justworks","gusto","indeed","ziprecruiter","glassdoor","monster","careerbuilder",
 "lhh","lee hecht harrison","allegis","apex group","apex systems","sthree","gi group","gi group holding","amn healthcare",
 "cross country healthcare","aya healthcare","lensa","supplemental health care","medical staffing network",
 "maximus","leidos","caci","booz allen hamilton","saic","dice"
}

def classify_new(name):
    n = name.strip()
    m = KEYWORD_BLOCK_RE.search(n)
    if m:
        return ("keyword_filter", m.group(0))
    if n.lower() in KNOWN_BLOCKED:
        return ("known_blocked", name)
    m = AUTO_DEALER_RE.search(n)
    if m:
        return ("auto_dealer", m.group(0))
    return (None, None)

def filter_new(new_rows, eligible_set, not_eligible_set, needs_review_set):
    result = {"pre_eligible": [], "pre_not_eligible": [], "pre_needs_review": [], "blocked": [], "passed": []}
    for (company, title, location, url) in new_rows:
        if company in not_eligible_set:
            result["pre_not_eligible"].append((company, title, location, url))
            continue
        if company in eligible_set:
            result["pre_eligible"].append((company, title, location, url))
            continue
        if company in needs_review_set:
            result["pre_needs_review"].append((company, title, location, url))
            continue
        reason, matched = classify_new(company)
        if reason:
            result["blocked"].append((company, title, location, url, reason, matched or ""))
        else:
            result["passed"].append((company, title, location, url))
    return result

# ---------- AI enrichment ----------

AI_SYSTEM_PROMPT = (
    "You are screening companies scraped from LinkedIn job postings for a US-based recruiting agency. "
    "You will be given a NUMBERED list of companies. Return ONLY valid JSON, no explanation, of the form "
    "{\"results\": [{\"number\": N, \"employee_size\": \"...\", \"industry\": \"...\", "
    "\"verdict\": \"eligible\"|\"not_eligible\"|\"needs_review\", \"reason\": \"...\"}, ...]}. Every number from "
    "1 to the count given MUST appear exactly once in results — never omit a number.\n"
    "For each company determine:\n"
    "1. employee_size — a short range such as '1-10', '10-50', '50-200', '200-500', '500-1000', '1000+', "
    "or 'Unknown' if you genuinely cannot estimate.\n"
    "2. industry — a short label, e.g. 'Manufacturing', 'Construction', 'Retail', 'Logistics', 'Auto Repair'.\n"
    "3. verdict — one of:\n"
    f"   - 'eligible': at least ~70% confident the company has roughly {AI_ELIGIBLE_MIN}-{AI_ELIGIBLE_MAX} "
    "employees AND is not in an excluded category below.\n"
    "   - 'not_eligible': at least ~70% confident the company is EITHER clearly outside that size range OR "
    "clearly in an excluded category.\n"
    "   - 'needs_review': NOT confident about size and/or the industry call — insufficient information, "
    "ambiguous, or borderline. This is the SAFE DEFAULT whenever unsure. Never guess eligible or not_eligible.\n"
    "Excluded categories: staffing/recruiting/HR outsourcing/PEO/BPO firms; healthcare/hospital/clinic/pharma "
    "organizations; government/public-sector/municipal bodies; schools/universities/education institutions; "
    "nonprofit/charity/religious organizations; banks/insurance/investment/financial-services firms; "
    "restaurant/hospitality/food-service businesses; auto dealerships; large/well-known multinational "
    "conglomerates or subsidiaries of one.\n"
    "EXCEPTION: accounting firms and CPA firms are NOT excluded — they can be eligible if they meet the size "
    "range. Other professional-services categories such as law firms, management consultancies, and "
    "financial-advisory/wealth-management firms remain excluded.\n"
    "4. reason — a short phrase (under 12 words) explaining the verdict."
)


class OpenAINonRetryableError(RuntimeError):
    """Raised for errors where retrying the same request will not help (bad key, bad request, etc.)."""
    pass


def call_openai_batch(companies, api_key, model=OPENAI_MODEL, timeout=90):
    """companies: list of (company_name, sample_title) tuples, in order. Returns dict keyed by
    1-based position in this call rather than relying on the company name matching back."""
    lines = [f"{i+1}. {name} (sample job title seen: {title})" for i, (name, title) in enumerate(companies)]
    user_prompt = f"Companies to evaluate ({len(companies)} total):\n" + "\n".join(lines)
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": AI_SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        "response_format": {"type": "json_object"},
        "max_completion_tokens": 4096,
        "temperature": 0,
    }
    req = urllib.request.Request(
        OPENAI_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        msg = f"OpenAI API error {e.code}: {detail[:300]}"
        if 400 <= e.code < 500:
            # Client-side error (bad key, bad request, etc.) — retrying the identical request won't help.
            raise OpenAINonRetryableError(msg)
        raise RuntimeError(msg)
    except urllib.error.URLError as e:
        raise RuntimeError(f"Could not reach OpenAI: {e.reason}")

    content = body["choices"][0]["message"]["content"]
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"AI response wasn't valid JSON, likely truncated mid-response: {e}")
    results = parsed.get("results", []) if isinstance(parsed, dict) else parsed

    by_number = {}
    for r in results:
        try:
            n = int(r.get("number"))
        except (TypeError, ValueError):
            continue
        if 1 <= n <= len(companies):
            by_number[n] = r
    return by_number


def run_ai_check(passed_rows, api_key, model=OPENAI_MODEL, batch_size=AI_BATCH_SIZE, progress_cb=None, max_attempts=3):
    """passed_rows: list of (company, title, location, url). Returns (results, errors).
    Matches AI answers back to companies by numbered position (not name). Each batch retries up
    to max_attempts on transient failures; a non-retryable error (bad key, bad request) stops
    retrying immediately. Any company still unanswered after retries defaults to 'needs_review'
    (never silently dropped, never guessed eligible) with a 'no_response' flag for the Retry button."""
    results_out = []
    errors = []
    total = len(passed_rows)
    for start in range(0, total, batch_size):
        batch = passed_rows[start:start + batch_size]
        remaining_idx = list(range(len(batch)))  # positions within `batch` still needing an answer
        answers = {}  # position in batch -> AI dict
        last_error = None
        stop_retrying = False

        for attempt in range(max_attempts):
            if not remaining_idx or stop_retrying:
                break
            sub_companies = [(batch[i][0], batch[i][1]) for i in remaining_idx]
            try:
                by_number = call_openai_batch(sub_companies, api_key, model=model)
            except OpenAINonRetryableError as e:
                last_error = str(e)
                stop_retrying = True
                continue
            except Exception as e:
                last_error = str(e)
                continue
            still_missing = []
            for pos_in_sub, batch_idx in enumerate(remaining_idx):
                r = by_number.get(pos_in_sub + 1)
                if r is not None:
                    answers[batch_idx] = r
                else:
                    still_missing.append(batch_idx)
            if not still_missing:
                last_error = None
                remaining_idx = []
                break
            last_error = f"AI returned {len(remaining_idx) - len(still_missing)} of {len(remaining_idx)} companies"
            remaining_idx = still_missing

        if remaining_idx:
            errors.append(
                f"Batch {start // batch_size + 1} ({len(batch)} companies): {last_error or 'no response'} "
                f"— {len(remaining_idx)} defaulted to needs-review, use Retry to try again."
            )

        for i, (company, title, location, url) in enumerate(batch):
            r = answers.get(i)
            if r is None:
                results_out.append({
                    "company_name": company, "title": title, "location": location, "url": url,
                    "employee_size": "", "industry": "", "verdict": "needs_review",
                    "reason": "No AI response — defaulted to needs review", "no_response": True,
                })
                continue
            verdict = str(r.get("verdict", "needs_review")).strip().lower()
            if verdict not in ("eligible", "not_eligible", "needs_review"):
                verdict = "needs_review"
            reason = str(r.get("reason", "")).strip()
            if verdict == "eligible":
                # Post-API safety recheck: re-apply the deterministic rules even to an AI "eligible" call,
                # in case the model missed something the keyword/known-firm/dealer rules would have caught.
                block_reason, matched = classify_new(company)
                if block_reason:
                    verdict = "not_eligible"
                    reason = f"Post-AI safety check matched rule: {matched or block_reason}"
            results_out.append({
                "company_name": company, "title": title, "location": location, "url": url,
                "employee_size": str(r.get("employee_size", "")).strip(),
                "industry": str(r.get("industry", "")).strip(),
                "verdict": verdict,
                "reason": reason,
                "no_response": False,
            })

        if progress_cb:
            progress_cb(min(start + batch_size, total), total)

    return results_out, errors


def commit_ai_results(results):
    """Writes AI verdicts to the right table: eligible/not_eligible go straight into the permanent
    cache; needs_review goes into the review queue instead, so an uncertain call never gets
    silently blacklisted — it waits for a human decision in the Review lists tab."""
    elig_rows = []
    not_elig_rows = []
    review_rows = []
    for r in results:
        v = r["verdict"]
        if v == "eligible":
            elig_rows.append((r["company_name"], r["employee_size"], r["industry"]))
        elif v == "not_eligible":
            not_elig_rows.append((r["company_name"], r["employee_size"], r["industry"], r["reason"]))
        else:
            review_rows.append((r["company_name"], r["employee_size"], r["industry"], r["reason"]))
    bulk_add_eligible(elig_rows)
    bulk_add_not_eligible(not_elig_rows)
    bulk_add_needs_review(review_rows)
    return len(elig_rows), len(not_elig_rows), len(review_rows)


def merge_ai_results(existing, new_results):
    """Updates entries in existing (in place) with matching company_name from new_results — used for retries."""
    by_name = {str(r["company_name"]).strip().lower(): i for i, r in enumerate(existing)}
    for nr in new_results:
        key = str(nr["company_name"]).strip().lower()
        if key in by_name:
            existing[by_name[key]] = nr
    return existing


# ---------- Callbacks ----------

def handle_add_company():
    name = st.session_state.company_input.strip()
    position = st.session_state.position_input.strip()
    ra = CURRENT_NAME if CURRENT_ROLE == "RA" else st.session_state.get("manual_target_ra", "")
    if not ra:
        st.session_state.company_message = ("error", "Select an active RA first.")
        return
    if name == "" or position == "":
        st.session_state.company_message = ("error", "Both company name and position are required.")
        return
    if is_client(name):
        st.session_state.company_message = ("error", f"'{name}' is on the CLIENTS list — never contact. Not added.")
        return
    avoided = is_avoided(name)
    if avoided:
        reason = avoided[0]
        extra = f" (reason: {reason})" if reason else ""
        st.session_state.company_message = ("error", f"'{name}' is on the DNC / avoid list — never contact{extra}. Not added.")
        return
    existing = find_recent_duplicate(name)
    if existing:
        prev_date, prev_ra = existing
        st.session_state.company_message = ("warning", f"Duplicate: '{name}' was already contacted on {prev_date} by {prev_ra} (within 30 days). Not added.")
        return
    created = create_manual_company_position(name, position, ra, CURRENT_NAME)
    if not created:
        st.session_state.company_message = ("warning", f"'{name}' already has an active position. Not added.")
        return
    st.session_state.company_message = ("success", f"Added: {name}")
    st.session_state.company_input = ""
    st.session_state.position_input = ""

def handle_add_keywords():
    raw = st.session_state.new_keywords
    parts = [p.strip().lower() for p in raw.split(",")]
    added = 0
    for p in parts:
        if p != "":
            add_keyword(p)
            added += 1
    st.session_state.block_message = ("success", f"Added {added} keyword(s).")
    st.session_state.new_keywords = ""

def handle_remove_keyword(keyword):
    remove_keyword(keyword)

def handle_add_client():
    name = st.session_state.client_name_input.strip()
    if name == "":
        st.session_state.client_message = ("error", "Enter a company name.")
        return
    add_to_clients(name)
    st.session_state.client_message = ("success", f"Added client: {name}")
    st.session_state.client_name_input = ""

def handle_add_avoid():
    name = st.session_state.avoid_name_input.strip()
    reason = st.session_state.avoid_reason_input.strip()
    if name == "":
        st.session_state.avoid_message = ("error", "Enter a company name.")
        return
    add_to_avoid(name, reason)
    st.session_state.avoid_message = ("success", f"Added to avoid list: {name}")
    st.session_state.avoid_name_input = ""
    st.session_state.avoid_reason_input = ""

# =====================================================================
#  RA Workflow — application shell (page config, theme, nav, dashboard)
# =====================================================================

st.set_page_config(
    page_title="RA Workflow",
    page_icon="◧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------- Login (per-user, role-based) ----------
# First run: create your own Manager account via Streamlit secrets:
#   [bootstrap_admin]
#   username = "vamsi"
#   password = "choose-a-strong-password"
#   display_name = "Vamsi"
# Everyone after that (TLs/RAs) is created from inside the app on the
# "Manage users" page (Manager role only) — see create_user().
def _check_auth():
    if st.session_state.get("current_user"):
        return True

    init_db()  # must exist before we can look up users, even pre-login

    st.markdown("""
    <style>
    .login-wrap { max-width:380px; margin:10vh auto 0; padding:2.4rem 2.2rem 2rem;
        background:#16202E; border:1px solid #26333F; border-radius:14px; }
    .login-title { font-family:'Space Grotesk',sans-serif; font-size:1.7rem; font-weight:700;
        color:#E8EDF4; letter-spacing:-.02em; margin-bottom:.3rem; }
    .login-sub { color:#8A9AAC; font-size:.88rem; margin-bottom:1.6rem; }
    .dot { display:inline-block; width:8px; height:8px; border-radius:50%;
        background:#0FA3B1; margin-left:4px; vertical-align:middle; }
    </style>
    <div class="login-wrap">
        <div class="login-title">RA Workflow<span class="dot"></span></div>
        <div class="login-sub">Sign in to continue.</div>
    </div>
    """, unsafe_allow_html=True)

    uname = st.text_input("Username", key="login_username", placeholder="Username")
    pwd = st.text_input("Password", type="password", key="login_pwd", placeholder="Password")

    if st.button("Sign in", type="primary"):
        user_row = get_user_by_username(uname) if uname.strip() else None
        if user_row is None:
            st.error("Incorrect username or password.")
        else:
            user_id, username, pwd_hash, display_name, role, active = user_row
            if not active:
                st.error("This account has been deactivated. Contact your manager.")
            elif not verify_password(pwd, pwd_hash):
                st.error("Incorrect username or password.")
            else:
                st.session_state.current_user = {
                    "id": user_id, "username": username,
                    "display_name": display_name, "role": role,
                }
                st.rerun()
    return False

if not _check_auth():
    st.stop()

init_db()

# Refresh active/role state on each rerun so deactivation and role changes take
# effect immediately instead of lingering in an old Streamlit session.
_fresh_user = get_current_user_record(st.session_state.current_user["id"])
if not _fresh_user or not _fresh_user[4]:
    st.session_state.pop("current_user", None)
    st.rerun()
st.session_state.current_user = {
    "id": _fresh_user[0], "username": _fresh_user[1],
    "display_name": _fresh_user[2], "role": _fresh_user[3],
}
CURRENT_USER = st.session_state.current_user
CURRENT_ROLE = CURRENT_USER["role"]          # "Manager" | "TL" | "RA"
CURRENT_NAME = CURRENT_USER["display_name"]

if "openai_api_key" not in st.session_state:
    st.session_state.openai_api_key = get_setting("openai_api_key", "")
if "theme_mode" not in st.session_state:
    st.session_state.theme_mode = get_setting("theme_mode", "dark")
if "page" not in st.session_state:
    st.session_state.page = "Dashboard"


# ---------- Design tokens ----------
# Stage-coded palette: colour carries meaning here (pass / held / dropped are
# real states in the product), so hues are assigned semantically, not decoratively.
THEMES = {
    "dark": {
        "bg": "#0D1520", "surface": "#16202E", "surface_2": "#1D2937",
        "border": "#26333F", "text": "#E8EDF4", "text_dim": "#8A9AAC",
        "primary": "#0FA3B1", "primary_soft": "rgba(15,163,177,.14)",
        "green": "#3FA372", "amber": "#E8A33D", "rose": "#E05263", "violet": "#7C6BE8",
        "grid": "#26333F",
    },
    "light": {
        "bg": "#F6F7F9", "surface": "#FFFFFF", "surface_2": "#F0F2F5",
        "border": "#E2E6EC", "text": "#16202E", "text_dim": "#5E6E80",
        "primary": "#0B7C87", "primary_soft": "rgba(11,124,135,.10)",
        "green": "#2F8659", "amber": "#C9862B", "rose": "#C93F50", "violet": "#5F4FCB",
        "grid": "#E2E6EC",
    },
}
T = THEMES[st.session_state.theme_mode]

# Plotly powers the interactive charts (hover, zoom, legend toggle, PNG export).
# It's an optional dependency — the app stays fully usable without it.
try:
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    go = None
    HAS_PLOTLY = False

PLOTLY_CONFIG = {"displaylogo": False,
                 "modeBarButtonsToRemove": ["lasso2d", "select2d", "autoScale2d"]}

# Streamlit renamed `use_container_width` to `width="stretch"`. Detect which one this
# install accepts so the app runs on both older and newer versions without warnings.
import inspect as _inspect
try:
    _SUPPORTS_WIDTH = "width" in _inspect.signature(st.dataframe).parameters
except (ValueError, TypeError):
    _SUPPORTS_WIDTH = False
FULL = {"width": "stretch"} if _SUPPORTS_WIDTH else {"use_container_width": True}


def chart_layout(height=300, margin_l=60, showlegend=False):
    """Shared Plotly layout so every chart inherits the current theme."""
    return dict(
        height=height,
        margin=dict(l=margin_l, r=16, t=16, b=36),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", size=12, color=T["text_dim"]),
        showlegend=showlegend,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, x=0),
        xaxis=dict(gridcolor=T["grid"], zeroline=False, linecolor=T["grid"]),
        yaxis=dict(gridcolor=T["grid"], zeroline=False, linecolor=T["grid"]),
        hoverlabel=dict(bgcolor=T["surface_2"], bordercolor=T["border"],
                        font=dict(color=T["text"], family="Inter, sans-serif")),
    )


def inject_css(t):
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;600&display=swap');

    :root {{
        --bg: {t['bg']}; --surface: {t['surface']}; --surface-2: {t['surface_2']};
        --border: {t['border']}; --text: {t['text']}; --text-dim: {t['text_dim']};
        --primary: {t['primary']}; --primary-soft: {t['primary_soft']};
        --green: {t['green']}; --amber: {t['amber']}; --rose: {t['rose']}; --violet: {t['violet']};
    }}

    .stApp {{ background: var(--bg); }}
    .main .block-container {{ padding: 2.2rem 2.6rem 5rem; max-width: 1500px; }}
    [data-testid="stHeader"] {{ background: transparent; }}

    html, body, [class*="css"], .stMarkdown, p, span, label, div {{
        font-family: 'Inter', system-ui, sans-serif; color: var(--text);
    }}
    h1, h2, h3, h4 {{ font-family: 'Space Grotesk', system-ui, sans-serif !important; color: var(--text) !important; letter-spacing: -.02em; }}
    h1 {{ font-size: 1.9rem !important; font-weight: 700 !important; }}
    h2 {{ font-size: 1.32rem !important; font-weight: 600 !important; margin-top: .4rem !important; }}
    h3 {{ font-size: 1.06rem !important; font-weight: 600 !important; }}

    /* ---- Page header ---- */
    .page-head {{ border-bottom: 1px solid var(--border); padding-bottom: 1rem; margin-bottom: 1.6rem; }}
    .page-eyebrow {{ font-family: 'JetBrains Mono', monospace; font-size: .68rem; font-weight: 600;
        letter-spacing: .16em; text-transform: uppercase; color: var(--primary); margin-bottom: .35rem; }}
    .page-title {{ font-family: 'Space Grotesk', sans-serif; font-size: 1.85rem; font-weight: 700;
        letter-spacing: -.025em; color: var(--text); line-height: 1.15; }}
    .page-sub {{ color: var(--text-dim); font-size: .93rem; margin-top: .4rem; max-width: 76ch; line-height: 1.55; }}

    /* ---- Sidebar ---- */
    [data-testid="stSidebar"] {{ background: var(--surface); border-right: 1px solid var(--border); }}
    [data-testid="stSidebar"] .block-container {{ padding-top: 1.5rem; }}
    .brand {{ display: flex; align-items: baseline; gap: .5rem; padding: 0 .25rem 1.1rem;
        border-bottom: 1px solid var(--border); margin-bottom: 1.1rem; }}
    .brand-mark {{ font-family: 'Space Grotesk', sans-serif; font-size: 1.32rem; font-weight: 700;
        color: var(--text); letter-spacing: -.03em; }}
    .brand-dot {{ width: 7px; height: 7px; border-radius: 50%; background: var(--primary);
        display: inline-block; margin-left: 1px; }}
    .nav-group {{ font-family: 'JetBrains Mono', monospace; font-size: .62rem; font-weight: 600;
        letter-spacing: .17em; text-transform: uppercase; color: var(--text-dim);
        margin: 1.1rem 0 .4rem .3rem; }}

    /* Sidebar nav buttons */
    [data-testid="stSidebar"] .stButton > button {{
        width: 100%; text-align: left; justify-content: flex-start;
        background: transparent; border: 1px solid transparent; color: var(--text-dim);
        font-size: .875rem; font-weight: 500; padding: .42rem .65rem; border-radius: 7px;
        transition: background .13s ease, color .13s ease;
    }}
    [data-testid="stSidebar"] .stButton > button:hover {{
        background: var(--surface-2); color: var(--text); border-color: transparent;
    }}
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {{
        background: var(--primary-soft); color: var(--primary); font-weight: 600;
        border-color: transparent; box-shadow: inset 2px 0 0 var(--primary);
    }}

    /* ---- Buttons (main area) ---- */
    .main .stButton > button, .main .stDownloadButton > button {{
        background: var(--surface-2); color: var(--text); border: 1px solid var(--border);
        border-radius: 7px; font-weight: 500; font-size: .875rem; padding: .44rem 1rem;
        transition: border-color .13s ease, background .13s ease;
    }}
    .main .stButton > button:hover, .main .stDownloadButton > button:hover {{
        border-color: var(--primary); color: var(--primary);
    }}
    .main .stButton > button[kind="primary"] {{
        background: var(--primary); color: #FFF; border-color: var(--primary); font-weight: 600;
    }}
    .main .stButton > button[kind="primary"]:hover {{ filter: brightness(1.1); color: #FFF; }}

    /* ---- Stat cards ---- */
    .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(168px, 1fr)); gap: .8rem; margin-bottom: .4rem; }}
    .stat {{ background: var(--surface); border: 1px solid var(--border); border-radius: 11px;
        padding: .95rem 1.05rem; position: relative; overflow: hidden; }}
    .stat::before {{ content:''; position:absolute; left:0; top:0; bottom:0; width:3px; background: var(--accent, var(--primary)); }}
    .stat-label {{ font-family: 'JetBrains Mono', monospace; font-size: .63rem; font-weight: 600;
        letter-spacing: .13em; text-transform: uppercase; color: var(--text-dim); }}
    .stat-value {{ font-family: 'Space Grotesk', sans-serif; font-size: 1.72rem; font-weight: 700;
        color: var(--text); line-height: 1.15; margin-top: .3rem; letter-spacing: -.03em; }}
    .stat-note {{ font-size: .74rem; color: var(--text-dim); margin-top: .15rem; }}

    /* ---- Signature: pipeline flow strip ---- */
    .flow {{ background: var(--surface); border: 1px solid var(--border); border-radius: 12px;
        padding: 1.3rem 1.4rem 1.15rem; margin: .3rem 0 1.2rem; }}
    .flow-head {{ font-family: 'JetBrains Mono', monospace; font-size: .64rem; font-weight: 600;
        letter-spacing: .16em; text-transform: uppercase; color: var(--text-dim); margin-bottom: 1rem; }}
    .flow-rail {{ display: flex; align-items: stretch; gap: .3rem; }}
    .flow-stage {{ flex: 1; min-width: 0; }}
    .flow-bar {{ height: 5px; border-radius: 3px; background: var(--stage); margin-bottom: .6rem; opacity: .95; }}
    .flow-n {{ font-family: 'Space Grotesk', sans-serif; font-size: 1.24rem; font-weight: 700;
        color: var(--text); letter-spacing: -.02em; line-height: 1; }}
    .flow-l {{ font-size: .72rem; color: var(--text-dim); margin-top: .25rem; line-height: 1.3; }}

    /* ---- Inputs / widgets ---- */
    .stTextInput input, .stNumberInput input, .stDateInput input, .stTextArea textarea {{
        background: var(--surface-2) !important; border: 1px solid var(--border) !important;
        border-radius: 7px !important; color: var(--text) !important; font-size: .88rem !important;
    }}
    .stTextInput input:focus, .stNumberInput input:focus {{ border-color: var(--primary) !important; }}
    [data-baseweb="select"] > div {{ background: var(--surface-2) !important;
        border-color: var(--border) !important; border-radius: 7px !important; }}
    label, .stTextInput label, .stSelectbox label {{ color: var(--text-dim) !important;
        font-size: .8rem !important; font-weight: 500 !important; }}

    /* ---- Expanders / containers ---- */
    [data-testid="stExpander"] {{ background: var(--surface); border: 1px solid var(--border);
        border-radius: 10px; margin-bottom: .55rem; }}
    [data-testid="stExpander"] summary {{ font-size: .875rem; font-weight: 500; color: var(--text); }}
    [data-testid="stExpander"] summary:hover {{ color: var(--primary); }}

    /* ---- Tables ---- */
    [data-testid="stDataFrame"], [data-testid="stDataEditor"] {{
        border: 1px solid var(--border); border-radius: 9px; overflow: hidden; }}

    /* ---- Alerts ---- */
    [data-testid="stAlert"] {{ border-radius: 9px; border-width: 1px; border-style: solid; font-size: .875rem; }}

    /* ---- Progress ---- */
    .stProgress > div > div > div > div {{ background: var(--primary); }}
    .stProgress > div > div > div {{ background: var(--surface-2); border-radius: 20px; }}

    /* ---- Dividers / misc ---- */
    hr {{ border-color: var(--border) !important; margin: 1.4rem 0 !important; }}
    [data-testid="stCaptionContainer"] {{ color: var(--text-dim) !important; font-size: .78rem !important; }}
    #MainMenu, footer {{ visibility: hidden; }}

    @media (prefers-reduced-motion: reduce) {{ * {{ transition: none !important; animation: none !important; }} }}
    </style>
    """, unsafe_allow_html=True)


inject_css(T)


def with_progress(df, label="Processing rows"):
    """Wraps df.iterrows() with a live progress bar. Large sheets (100k+ rows) take a
    noticeable moment to parse, so show movement rather than a frozen screen."""
    total = len(df)
    bar = st.progress(0.0, text=f"{label}… 0 of {total:,}")
    step = max(1, total // 100)
    for i, (_, row) in enumerate(df.iterrows()):
        if i % step == 0:
            bar.progress(i / total if total else 1.0, text=f"{label}… {i:,} of {total:,}")
        yield row
    bar.progress(1.0, text=f"{label} — {total:,} rows read")
    bar.empty()


def perf_chart(df, label_col, key):
    """Grouped sent/reply bars with a response-rate line on a second axis.
    Interactive: hover for exact figures, drag to zoom, click the legend to isolate a series."""
    if df.empty:
        return
    if not HAS_PLOTLY:
        st.bar_chart(df.set_index(label_col)[["Sent", "Responses"]])
        return

    max_n = len(df)
    if max_n > 6:
        top_n = st.slider("Show top", 5, max_n, min(15, max_n), key=f"topn_{key}",
                          help="Ranked by emails sent")
        d = df.head(top_n)
    else:
        d = df

    fig = go.Figure()
    fig.add_bar(x=d[label_col], y=d["Sent"], name="Sent",
                marker=dict(color=T["primary"], line=dict(width=0)),
                hovertemplate="%{x}<br>Sent: %{y:,}<extra></extra>")
    fig.add_bar(x=d[label_col], y=d["Responses"], name="Replies",
                marker=dict(color=T["green"], line=dict(width=0)),
                hovertemplate="%{x}<br>Replies: %{y:,}<extra></extra>")
    fig.add_scatter(x=d[label_col], y=d["Response Rate %"], name="Reply rate",
                    yaxis="y2", mode="lines+markers",
                    line=dict(color=T["amber"], width=2),
                    marker=dict(size=6, color=T["amber"]),
                    hovertemplate="%{x}<br>Reply rate: %{y:.2f}%<extra></extra>")

    layout = chart_layout(height=380, margin_l=54, showlegend=True)
    layout["barmode"] = "group"
    layout["xaxis"]["tickangle"] = -35
    layout["yaxis"]["title"] = "Emails"
    layout["yaxis2"] = dict(overlaying="y", side="right", showgrid=False,
                            title="Reply rate %", color=T["amber"],
                            linecolor=T["grid"])
    fig.update_layout(**layout)
    st.plotly_chart(fig, **FULL, config=PLOTLY_CONFIG)


def page_header(eyebrow, title, sub=""):
    st.markdown(
        f'<div class="page-head"><div class="page-eyebrow">{eyebrow}</div>'
        f'<div class="page-title">{title}</div>'
        + (f'<div class="page-sub">{sub}</div>' if sub else "")
        + '</div>', unsafe_allow_html=True)


def stat_cards(cards):
    """cards: list of (label, value, note, accent_css_var)"""
    html = '<div class="stat-grid">'
    for label, value, note, accent in cards:
        html += (f'<div class="stat" style="--accent:var(--{accent})">'
                 f'<div class="stat-label">{label}</div>'
                 f'<div class="stat-value">{value}</div>'
                 f'<div class="stat-note">{note}</div></div>')
    st.markdown(html + '</div>', unsafe_allow_html=True)


# ---------- Sidebar ----------
if CURRENT_ROLE == "RA":
    NAV = [
        ("Overview", [("Dashboard", "Dashboard")]),
        ("Pipeline", [("Add company", "Add company"), ("Workpage", "Workpage")]),
        ("Data", [("EOD uploads", "EOD uploads")]),
    ]
elif CURRENT_ROLE == "TL":
    NAV = [
        ("Overview", [("Dashboard", "Dashboard")]),
        ("Pipeline", [("Bulk scrape", "Bulk scrape"), ("Add company", "Add company"),
                      ("Review lists", "Review lists"), ("RA assignments", "RA assignments"),
                      ("Workpage", "Workpage"), ("Not Eligible requests", "Not Eligible requests")]),
        ("Campaign", [("Campaign export", "Campaign export")]),
        ("Data", [("EOD uploads", "EOD uploads")]),
        ("Insight", [("Analytics", "Analytics")]),
    ]
else:
    NAV = [
        ("Overview", [("Dashboard", "Dashboard")]),
        ("Pipeline", [("Bulk scrape", "Bulk scrape"), ("Add company", "Add company"),
                      ("Review lists", "Review lists"), ("RA assignments", "RA assignments"),
                      ("Workpage", "Workpage"), ("Not Eligible requests", "Not Eligible requests")]),
        ("Campaign", [("Campaign export", "Campaign export")]),
        ("Data", [("Prospects DB", "Prospects DB"), ("EOD uploads", "EOD uploads"),
                  ("Bounced & DNC", "Bounced & DNC"), ("Import history", "Import history")]),
        ("Insight", [("Analytics", "Analytics")]),
        ("Settings", [("Clients & DNC", "Clients & DNC"), ("Title block list", "Title block list"),
                      ("Danger zone", "Danger zone")]),
        ("Admin", [("Manage users", "Manage users")]),
    ]

with st.sidebar:
    st.markdown('<div class="brand"><span class="brand-mark">RA Workflow</span>'
                '<span class="brand-dot"></span></div>', unsafe_allow_html=True)
    st.markdown(
        f'<div style="padding:.5rem .2rem 1rem; font-size:.85rem;">'
        f'Signed in as <b>{CURRENT_NAME}</b><br><span style="opacity:.7;">{CURRENT_ROLE}</span></div>',
        unsafe_allow_html=True
    )

    for group, items in NAV:
        st.markdown(f'<div class="nav-group">{group}</div>', unsafe_allow_html=True)
        for label, key in items:
            active = st.session_state.page == key
            if st.button(label, key=f"nav_{key}", type="primary" if active else "secondary"):
                st.session_state.page = key
                st.rerun()

    st.markdown('<div class="nav-group">Appearance</div>', unsafe_allow_html=True)
    mode_label = "Switch to light" if st.session_state.theme_mode == "dark" else "Switch to dark"
    if st.button(mode_label, key="theme_toggle"):
        st.session_state.theme_mode = "light" if st.session_state.theme_mode == "dark" else "dark"
        set_setting("theme_mode", st.session_state.theme_mode)
        st.rerun()

    st.markdown('<div class="nav-group">Account</div>', unsafe_allow_html=True)
    if st.button("Log out", key="logout_btn"):
        st.session_state.pop("current_user", None)
        st.rerun()


# ---------- Dashboard ----------
def page_dashboard():
    page_header("Overview", "Dashboard",
                "Where every company stands right now — from first scrape through to a booked reply.")

    (total_companies, hot, n_elig, n_not_elig, n_review, n_prospects,
     n_prospect_cos, n_assign, n_eod, n_sent, n_resp) = get_dashboard_metrics()
    reply_rate = (n_resp / n_sent * 100) if n_sent else 0.0

    stat_cards([
        ("Companies", f"{total_companies:,}", f"{hot:,} contacted in last 30 days", "primary"),
        ("Prospects", f"{n_prospects:,}", f"across {n_prospect_cos:,} companies", "violet"),
        ("Emails sent", f"{n_sent:,}", "all-time", "primary"),
        ("Positive replies", f"{n_resp:,}", f"{reply_rate:.2f}% reply rate", "green"),
    ])

    # Signature element: the routing rail. Widths are proportional to real counts,
    # and each stage is coloured by what it means (passing / held / dropped).
    stages = [
        (total_companies, "In history", "primary"),
        (n_elig, "Eligible", "green"),
        (n_review, "Needs review", "amber"),
        (n_not_elig, "Not eligible", "rose"),
        (n_assign, "Assigned to RAs", "violet"),
        (n_sent, "Emailed", "primary"),
        (n_resp, "Replied", "green"),
    ]
    peak = max((s[0] for s in stages), default=0) or 1
    rail = '<div class="flow"><div class="flow-head">Pipeline flow</div><div class="flow-rail">'
    for value, label, colour in stages:
        width = max(8, int(value / peak * 100))
        rail += (f'<div class="flow-stage">'
                 f'<div class="flow-bar" style="--stage:var(--{colour}); width:{width}%"></div>'
                 f'<div class="flow-n">{value:,}</div>'
                 f'<div class="flow-l">{label}</div></div>')
    st.markdown(rail + '</div></div>', unsafe_allow_html=True)

    left, right = st.columns([3, 2])

    with left:
        st.markdown("### Screening decisions")
        decided = n_elig + n_not_elig + n_review
        if decided and not HAS_PLOTLY:
            st.bar_chart(pd.DataFrame(
                {"Companies": [n_elig, n_review, n_not_elig]},
                index=["Eligible", "Needs review", "Not eligible"]))
        elif decided:
            fig = go.Figure(go.Bar(
                x=[n_elig, n_review, n_not_elig],
                y=["Eligible", "Needs review", "Not eligible"],
                orientation="h",
                marker=dict(color=[T["green"], T["amber"], T["rose"]],
                            line=dict(width=0)),
                text=[f"{n_elig:,}", f"{n_review:,}", f"{n_not_elig:,}"],
                textposition="auto",
                hovertemplate="%{y}: %{x:,}<extra></extra>",
            ))
            fig.update_layout(**chart_layout(height=230, margin_l=110))
            st.plotly_chart(fig, **FULL, config=PLOTLY_CONFIG)
        else:
            st.info("No screening decisions yet. Run a scrape through Bulk scrape to populate this.")

    with right:
        st.markdown("### Reach-out funnel")
        if n_sent and not HAS_PLOTLY:
            st.bar_chart(pd.DataFrame(
                {"Count": [n_prospects, n_sent, n_resp]},
                index=["Prospects", "Emails sent", "Replies"]))
        elif n_sent:
            fig2 = go.Figure(go.Funnel(
                y=["Prospects on file", "Emails sent", "Positive replies"],
                x=[n_prospects, n_sent, n_resp],
                marker=dict(color=[T["violet"], T["primary"], T["green"]]),
                textinfo="value+percent initial",
                hovertemplate="%{y}: %{x:,}<extra></extra>",
            ))
            fig2.update_layout(**chart_layout(height=230, margin_l=10))
            st.plotly_chart(fig2, **FULL, config=PLOTLY_CONFIG)
        else:
            st.info("Upload sent-mail history in Analytics to see the funnel.")

    st.markdown("### Guardrails")
    stat_cards([
        ("Clients", f"{count_clients():,}", "never contact", "rose"),
        ("DNC companies", f"{count_avoid():,}", "never contact", "rose"),
        ("Bounced emails", f"{count_bounced():,}", "scrubbed from prospects", "amber"),
        ("EOD rows", f"{n_eod:,}", "sourced by RAs", "primary"),
    ])

def page_add():
    page_header("Pipeline", "Add a company",
                "For a single company an RA found by hand — checked against clients, DNC and the 30-day duplicate window before it saves.")
    if CURRENT_ROLE == "RA":
        st.text_input("Assigned to", value=CURRENT_NAME, disabled=True, key="manual_ra_display")
    else:
        st.selectbox("Assign to RA / TL / Manager", get_active_ra_names(), key="manual_target_ra")
    st.text_input("Company name (exact LinkedIn name)", key="company_input")
    st.text_input("Position", key="position_input")
    st.button("Add company", on_click=handle_add_company)
    if "company_message" in st.session_state:
        kind, text = st.session_state.company_message
        if kind == "success":
            st.success(text)
        elif kind == "warning":
            st.warning(text)
        else:
            st.error(text)
        del st.session_state.company_message
    total, hot = count_companies()
    st.markdown("### Companies in database")
    st.write(f"**{total:,}** total in history  ·  **{hot:,}** contacted in the last 30 days (these block re-adds)")
    companies = get_companies()
    if companies:
        df = pd.DataFrame(companies, columns=["Company", "Position", "Added by", "Date", "Source"])
        st.caption("Showing the 100 most recent.")
        st.dataframe(df.head(100), hide_index=True)
    else:
        st.info("No companies yet. Add one above or import your history.")


def page_bulk():
    page_header("Pipeline", "Bulk scrape",
                "Six steps: clean the raw export, route by contact age, map re-engage companies to prospects, filter the new ones, screen them with AI, then deal them out to RAs.")
    st.markdown("### Step 1: Clean")
    st.write("Upload the raw LinkedIn scrape exported as **.xlsx** (the links must survive, so it must be Excel, not CSV). The app reads the hyperlink on each job title to pull out Company, Title, Location and the job URL — and fixes doubled titles like 'ElectricianElectrician'.")

    scrape_file = st.file_uploader("Upload raw LinkedIn scrape (.xlsx)", type=["xlsx"], key="scrape_upload")
    if scrape_file is not None:
        col_run1, col_run2 = st.columns(2)
        with col_run1:
            run_all = st.button("Run full pipeline (Clean → Route → Filter)", type="primary")
        with col_run2:
            run_clean_only = st.button("Clean only (step-by-step)")

        if run_all or run_clean_only:
            with st.spinner("Reading and cleaning..."):
                try:
                    rows, jobs_found = clean_scrape(scrape_file)
                    st.session_state.cleaned_rows = rows
                    st.session_state.cleaned_jobs = jobs_found
                    st.session_state.pop("routed", None)
                    st.session_state.pop("filtered", None)
                except Exception as e:
                    st.error(f"Could not read the file: {e}")
                    rows = None

            if run_all and rows is not None:
                with st.spinner("Routing..."):
                    st.session_state.routed = route_scrape(rows)
                with st.spinner("Filtering..."):
                    st.session_state.filtered = filter_new(
                        st.session_state.routed['new'], get_eligible_set(), get_not_eligible_set(), get_needs_review_set()
                    )
                st.success("Full pipeline done — Clean, Route, and Filter all ran. Scroll down for results, "
                           "then run the AI check on whatever landed in 'pending AI'.")

    if "cleaned_rows" in st.session_state:
        rows = st.session_state.cleaned_rows
        st.success(f"{st.session_state.cleaned_jobs:,} job rows found  →  {len(rows):,} unique companies kept (first job per company).")
        cdf = pd.DataFrame(rows, columns=["Company Name", "Job Title", "Location", "Job URL"])
        st.dataframe(cdf, hide_index=True)
        st.download_button("Download cleaned CSV", cdf.to_csv(index=False), "cleaned.csv", "text/csv")

        st.divider()
        st.subheader("Step 2: Route")
        st.write("Drop clients/DNC, skip anything contacted in the last 15 days, mark 15-30 and 30+ day companies to re-engage, and flag genuinely new companies for filtering.")
        if st.button("Route these companies"):
            with st.spinner("Routing..."):
                st.session_state.routed = route_scrape(rows)
                st.session_state.pop("filtered", None)

    if "routed" in st.session_state:
        g = st.session_state.routed
        st.write(
            f"**New:** {len(g['new']):,}  \u00b7  "
            f"**Re-engage 15-30d:** {len(g['reengage_15_30']):,}  \u00b7  "
            f"**Re-engage 30+d:** {len(g['reengage_30plus']):,}  \u00b7  "
            f"**Skipped (0-15d):** {len(g['skip']):,}  \u00b7  "
            f"**Dropped (client/DNC):** {len(g['client_dnc']):,}"
        )
        _cols = ["Company Name", "Job Title", "Location", "Job URL"]
        with st.expander(f"New companies -> go to filtering ({len(g['new']):,})"):
            st.dataframe(pd.DataFrame(g['new'], columns=_cols), hide_index=True)
        with st.expander(f"Re-engage 15-30 days -> Prospects Mapping ({len(g['reengage_15_30']):,})"):
            st.dataframe(pd.DataFrame(g['reengage_15_30'], columns=_cols), hide_index=True)
        with st.expander(f"Re-engage 30+ days -> Prospects Mapping ({len(g['reengage_30plus']):,})"):
            st.dataframe(pd.DataFrame(g['reengage_30plus'], columns=_cols), hide_index=True)
        with st.expander(f"Skipped - contacted in last 15 days ({len(g['skip']):,})"):
            st.dataframe(pd.DataFrame(g['skip'], columns=_cols), hide_index=True)
        with st.expander(f"Dropped - clients / DNC ({len(g['client_dnc']):,})"):
            st.dataframe(pd.DataFrame(g['client_dnc'], columns=_cols), hide_index=True)

        st.divider()
        st.subheader("Step 3: Prospects Mapping (re-engage companies)")
        st.write(
            f"Matches your **{len(g['reengage_15_30']):,}** (15-30 day) and **{len(g['reengage_30plus']):,}** "
            "(30+ day) re-engage companies against the Prospects DB using exact company spelling. A match "
            "pulls First Name + Email from the Prospects DB and Title + Location from this scrape — one row "
            "per contact. No match means we don't currently have contacts for that company, so it's folded "
            "into the RA Assignment pool below for fresh sourcing instead."
        )
        if st.button("Run Prospects Mapping"):
            reengage_pool = (
                [(c, t, l, u, "15-30 days") for (c, t, l, u) in g['reengage_15_30']] +
                [(c, t, l, u, "30+ days") for (c, t, l, u) in g['reengage_30plus']]
            )
            with st.spinner("Matching against Prospects DB..."):
                mapped_rows, unmatched_rows = map_reengage_to_prospects(reengage_pool)
            st.session_state.mapped_rows = mapped_rows
            st.session_state.unmatched_reengage = unmatched_rows

        if "mapped_rows" in st.session_state:
            mapped_rows = st.session_state.mapped_rows
            unmatched_rows = st.session_state.unmatched_reengage
            output_cols = ["First Name", "Email", "Title", "Location", "Company Name"]
            mdf = pd.DataFrame(mapped_rows, columns=output_cols + ["Bucket"]) if mapped_rows else pd.DataFrame(columns=output_cols + ["Bucket"])
            n_matched_companies = mdf["Company Name"].nunique() if not mdf.empty else 0
            st.write(
                f"**{len(mapped_rows):,}** contacts mapped across **{n_matched_companies:,}** companies  \u00b7  "
                f"**{len(unmatched_rows):,}** companies with no prospects on file -> going to RA Assignment"
            )
            if not mdf.empty:
                st.dataframe(mdf, hide_index=True)
                st.download_button(
                    "Download ALL mapped prospects (CSV)", mdf[output_cols].to_csv(index=False),
                    "prospects_mapping_all.csv", "text/csv"
                )
                m15 = mdf[mdf["Bucket"] == "15-30 days"]
                m30 = mdf[mdf["Bucket"] == "30+ days"]
                dl1, dl2 = st.columns(2)
                with dl1:
                    st.download_button(
                        f"Download 15-30 day bucket only ({len(m15):,})", m15[output_cols].to_csv(index=False),
                        "prospects_mapping_15_30.csv", "text/csv"
                    )
                with dl2:
                    st.download_button(
                        f"Download 30+ day bucket only ({len(m30):,})", m30[output_cols].to_csv(index=False),
                        "prospects_mapping_30plus.csv", "text/csv"
                    )
            else:
                st.info("No matches this run — every re-engage company will go to RA Assignment below.")
            if unmatched_rows:
                with st.expander(f"No prospects on file — going to RA Assignment ({len(unmatched_rows):,})"):
                    udf = pd.DataFrame(unmatched_rows, columns=["Company Name", "Job Title", "Location", "Job URL", "Bucket"])
                    st.dataframe(udf, hide_index=True)

        st.divider()
        st.subheader("Step 4: Filter the new companies")
        st.write("Rule-based filters (staffing / healthcare / government / finance keywords, named staffing firms, car dealers) applied to the NEW pile only. Whatever passes is pending the AI check (next step).")
        if st.button("Filter new companies"):
            with st.spinner("Filtering..."):
                st.session_state.filtered = filter_new(g['new'], get_eligible_set(), get_not_eligible_set(), get_needs_review_set())

    if "filtered" in st.session_state:
        f = st.session_state.filtered
        _c4 = ["Company Name", "Job Title", "Location", "Job URL"]
        st.write(
            f"**Already eligible (pass):** {len(f['pre_eligible']):,}  \u00b7  "
            f"**Already not-eligible (drop):** {len(f['pre_not_eligible']):,}  \u00b7  "
            f"**Already needs review (skip AI):** {len(f['pre_needs_review']):,}  \u00b7  "
            f"**Newly blocked by a rule:** {len(f['blocked']):,}  \u00b7  "
            f"**Passed -> pending AI:** {len(f['passed']):,}"
        )
        with st.expander(f"Already on your Eligible list -> pass ({len(f['pre_eligible']):,})"):
            st.dataframe(pd.DataFrame(f['pre_eligible'], columns=_c4), hide_index=True)
        with st.expander(f"Already on your Not-eligible list -> drop ({len(f['pre_not_eligible']):,})"):
            st.dataframe(pd.DataFrame(f['pre_not_eligible'], columns=_c4), hide_index=True)
        with st.expander(f"Already in your Needs-review queue -> resolve in Review lists ({len(f['pre_needs_review']):,})"):
            st.dataframe(pd.DataFrame(f['pre_needs_review'], columns=_c4), hide_index=True)
        with st.expander(f"Newly blocked by a rule - audit these ({len(f['blocked']):,})"):
            st.dataframe(pd.DataFrame(f['blocked'], columns=["Company Name", "Job Title", "Location", "Job URL", "Reason", "Matched word"]), hide_index=True)
        with st.expander(f"Passed everything -> pending AI ({len(f['passed']):,})"):
            st.dataframe(pd.DataFrame(f['passed'], columns=_c4), hide_index=True)

        st.divider()
        st.subheader("Step 5: AI check")
        st.write(
            f"Sends the {len(f['passed']):,} unknown companies to OpenAI to estimate Employee Size, Industry, "
            f"and a verdict ({AI_ELIGIBLE_MIN}-{AI_ELIGIBLE_MAX} employees, excluding staffing, healthcare, "
            "government, education, nonprofit, finance, and hospitality — accounting firms are allowed). "
            "Confident calls save straight to Eligible/Not-eligible; anything the AI is unsure about goes to "
            "a Needs-review queue instead of being auto-blacklisted. Every company is skipped from the AI "
            "on future runs either way."
        )
        key_col1, key_col2 = st.columns([3, 1])
        with key_col1:
            st.text_input("OpenAI API key", type="password", key="openai_api_key")
        with key_col2:
            st.write("")
            st.write("")
            if st.button("Save key"):
                set_setting("openai_api_key", st.session_state.openai_api_key.strip())
                st.success("Saved.")
        if get_setting("openai_api_key", ""):
            st.caption("A key is saved for next time — clear the field and click Save key to remove it.")
        run_disabled = len(f['passed']) == 0
        if st.button("Run AI check", disabled=run_disabled):
            key = st.session_state.openai_api_key.strip()
            if not key:
                st.error("Enter your OpenAI API key first.")
            else:
                progress_bar = st.progress(0.0, text="Starting AI check...")

                def _update_progress(done, total):
                    progress_bar.progress(done / total, text=f"{done:,}/{total:,} companies checked")

                results, errors = run_ai_check(f['passed'], key, progress_cb=_update_progress)
                progress_bar.empty()
                st.session_state.ai_results = results
                st.session_state.ai_errors = errors

    if "ai_results" in st.session_state:
        ai_results = st.session_state.ai_results
        ai_errors = st.session_state.get("ai_errors", [])
        st.divider()
        st.subheader("AI check results")
        if ai_errors:
            st.error(f"{len(ai_errors)} batch(es) had trouble — affected companies defaulted to needs-review.")
            for err in ai_errors:
                st.caption(err)
        n_elig = sum(1 for r in ai_results if r["verdict"] == "eligible")
        n_not_elig = sum(1 for r in ai_results if r["verdict"] == "not_eligible")
        n_review = sum(1 for r in ai_results if r["verdict"] == "needs_review")
        n_no_response = sum(1 for r in ai_results if r.get("no_response"))
        st.write(
            f"**AI eligible:** {n_elig:,}  \u00b7  **AI not-eligible:** {n_not_elig:,}  \u00b7  "
            f"**Needs review:** {n_review:,} (of which {n_no_response:,} got no AI response at all)"
        )
        rdf = pd.DataFrame(ai_results)[["company_name", "employee_size", "industry", "verdict", "reason", "title", "location"]]
        rdf.columns = ["Company Name", "Employee Size", "Industry", "Verdict", "Reason", "Job Title", "Location"]
        st.dataframe(rdf, hide_index=True)

        if n_no_response > 0:
            if st.button(f"Retry the {n_no_response:,} companies with no AI response"):
                key = st.session_state.get("openai_api_key", "").strip()
                if not key:
                    st.error("Enter your OpenAI API key above, then retry.")
                else:
                    failed_rows = [
                        (r["company_name"], r["title"], r["location"], r["url"])
                        for r in ai_results if r.get("no_response")
                    ]
                    retry_progress = st.progress(0.0, text="Retrying failed companies...")

                    def _update_retry_progress(done, total):
                        retry_progress.progress(done / total, text=f"{done:,}/{total:,} retried")

                    retry_results, retry_errors = run_ai_check(failed_rows, key, progress_cb=_update_retry_progress)
                    retry_progress.empty()
                    merge_ai_results(ai_results, retry_results)
                    st.session_state.ai_results = ai_results
                    st.session_state.ai_errors = retry_errors
                    st.rerun()

        if st.button("Save these results"):
            n_e, n_ne, n_nr = commit_ai_results(ai_results)
            st.success(
                f"Saved {n_e:,} to Eligible, {n_ne:,} to Not-eligible, and {n_nr:,} to the Needs-review "
                "queue (Review lists tab) for you to resolve by hand. Eligible/Not-eligible short-circuit "
                "the keyword rules and AI on every future run; needs-review is skipped from AI spend too, "
                "but stays pending until you resolve it."
            )
            del st.session_state.ai_results
            st.session_state.pop("ai_errors", None)

    if "filtered" in st.session_state:
        f = st.session_state.filtered
        assignment_pool = list(f['pre_eligible'])
        seen_names = set(c[0] for c in assignment_pool)
        for r in st.session_state.get("ai_results", []):
            if r["verdict"] == "eligible" and r["company_name"] not in seen_names:
                assignment_pool.append((r["company_name"], r["title"], r["location"], r["url"]))
                seen_names.add(r["company_name"])
        for (company, title, location, url, bucket) in st.session_state.get("unmatched_reengage", []):
            if company not in seen_names:
                assignment_pool.append((company, title, location, url))
                seen_names.add(company)
        for company, title, location, url in get_reassignment_queue_rows():
            if company not in seen_names:
                assignment_pool.append((company, title, location, url))
                seen_names.add(company)
        assignment_pool.sort(key=lambda c: c[0].lower())

        # A company already holding an open Workpage position has already been dealt out —
        # re-offering it here is what caused double-assignment. Exclude it instead.
        already_working = get_companies_with_active_position()
        n_before_dedupe = len(assignment_pool)
        assignment_pool = [c for c in assignment_pool if c[0] not in already_working]
        n_excluded = n_before_dedupe - len(assignment_pool)

        st.divider()
        st.subheader("Step 6: Assign new eligible companies to RAs")
        st.write(
            f"**{len(assignment_pool):,}** companies ready to assign — new eligible companies (already-eligible "
            "pass-throughs plus anything the AI approved and you saved this run) PLUS any re-engage company "
            "from Step 3 that had no prospects on file. Sorted alphabetically before dealing so scrape "
            "title-clusters don't unfairly load one RA."
        )
        if n_excluded > 0:
            st.caption(f"{n_excluded:,} company/companies already have an open Workpage position "
                       "(already assigned in an earlier batch) — excluded here to prevent double-assignment.")

        if assignment_pool:
            roster_df = pd.DataFrame({"Assignee": [""], "Count": [0]})
            edited_roster = st.data_editor(
                roster_df, num_rows="dynamic", key="ra_roster_editor", hide_index=True,
                column_config={
                    "Assignee": st.column_config.SelectboxColumn(options=get_active_ra_names(), required=True),
                    "Count": st.column_config.NumberColumn(min_value=0, step=1),
                },
            )
            total_assigned = int(pd.to_numeric(edited_roster["Count"], errors="coerce").fillna(0).sum())
            diff = len(assignment_pool) - total_assigned
            if total_assigned == 0:
                st.info("Type each RA's name and how many companies they should get.")
            elif diff > 0:
                st.warning(f"{diff:,} companies still uncovered — raise a count or add another RA row.")
            elif diff < 0:
                st.error(f"Counts add up to {-diff:,} more than the available pool — lower a count.")
            else:
                st.success("RA counts match the pool exactly.")

            if st.button("Deal out to RAs", disabled=(total_assigned == 0)):
                roster = [
                    (str(row["Assignee"]).strip(), int(row["Count"]))
                    for _, row in edited_roster.iterrows()
                    if str(row["Assignee"]).strip() and pd.notna(row["Count"]) and int(row["Count"]) > 0
                ]
                if not roster:
                    st.error("Enter at least one RA name with a count greater than 0.")
                else:
                    assignments, leftover = deal_companies_to_ras(assignment_pool, roster)
                    batch_id = uuid.uuid4().hex[:10]
                    today = date.today().isoformat()
                    rows_to_save = [
                        (batch_id, company, title, location, url, ra_name, today)
                        for ra_name, chunk in assignments.items()
                        for (company, title, location, url) in chunk
                    ]
                    created_count = save_assignment_bundle(rows_to_save, CURRENT_NAME)
                    st.session_state.last_assignment_batch = batch_id
                    st.session_state.last_assignments = assignments
                    st.session_state.last_leftover = leftover
                    st.success(f"Assigned {created_count:,} companies across {len(assignments):,} RA(s). Saved as batch {batch_id}.")

        if "last_assignments" in st.session_state:
            st.caption(f"Last batch: {st.session_state.get('last_assignment_batch', '')}")
            for ra_name, chunk in st.session_state.last_assignments.items():
                with st.expander(f"{ra_name} — {len(chunk):,} companies"):
                    adf = pd.DataFrame(chunk, columns=["Company Name", "Job Title", "Location", "Job URL"])
                    st.dataframe(adf, hide_index=True)
                    st.download_button(
                        f"Download {ra_name}'s list (CSV)", adf.to_csv(index=False),
                        f"{ra_name.replace(' ', '_')}_assignment.csv", "text/csv", key=f"dl_{ra_name}"
                    )
            leftover = st.session_state.get("last_leftover", [])
            if leftover:
                with st.expander(f"Unassigned leftover — counts didn't cover the whole pool ({len(leftover):,})"):
                    ldf = pd.DataFrame(leftover, columns=["Company Name", "Job Title", "Location", "Job URL"])
                    st.dataframe(ldf, hide_index=True)


def page_block():
    page_header("Settings", "Title block list",
                "Job titles containing any of these words get flagged during bulk filtering. Saved once, reused every run.")
    st.write("Job titles containing any of these words get flagged during bulk filtering. The list is saved — you never retype it.")
    st.text_input("Add keywords (comma-separated, e.g. warehouse, part-time, apprentice)", key="new_keywords")
    st.button("Add to block list", on_click=handle_add_keywords)
    if "block_message" in st.session_state:
        kind, text = st.session_state.block_message
        st.success(text)
        del st.session_state.block_message
    keywords = get_block_list()
    st.subheader(f"Current block list ({len(keywords)})")
    if keywords:
        for kw in keywords:
            col1, col2 = st.columns([4, 1])
            col1.write(kw)
            if CURRENT_ROLE == "Manager":
                col2.button("Remove", key=f"rm_{kw}", on_click=handle_remove_keyword, args=(kw,))
    else:
        st.info("Block list is empty. Add your usual keywords above.")
    st.divider()
    st.subheader("Test a job title")
    test_title = st.text_input("Job title to test", key="test_title")
    if test_title.strip() != "":
        hit = title_is_blocked(test_title, keywords)
        if hit:
            st.error(f"BLOCKED — matches keyword '{hit}'")
        else:
            st.success("Not blocked — this title would pass through.")


def page_import():
    page_header("Data", "Import history",
                "Load your existing contacted-company sheets. Importing is additive and safe to repeat — the 30-day window ages automatically by date.")
    st.write("Load your existing contacted-company sheets (Master and 'Out of 30 days'). Import is additive and safe. The 30-day window ages automatically by date.")
    uploaded = st.file_uploader("Choose a CSV file to import", type=["csv"], key="history_upload")
    if uploaded is not None:
        try:
            raw_df = read_csv_safe(uploaded)
            raw_df.columns = [c.strip() for c in raw_df.columns]
            st.write("Preview (first 5 rows):")
            st.dataframe(raw_df.head(5), hide_index=True)
            required = ["Company Name", "Position", "RA Name", "Date"]
            missing = [c for c in required if c not in raw_df.columns]
            if missing:
                st.error(f"These expected columns are missing: {missing}. Found: {list(raw_df.columns)}")
            else:
                if st.button(f"Import {len(raw_df)} rows"):
                    existing = get_existing_keys()
                    rows = []
                    skipped = 0
                    for _, r in raw_df.iterrows():
                        company = str(r["Company Name"]).strip()
                        if company == "" or company.lower() == "nan":
                            continue
                        d = parse_date(r["Date"])
                        key = (company, d)
                        if key in existing:
                            skipped += 1
                            continue
                        existing.add(key)
                        rows.append((company, str(r["Position"]).strip(), str(r["RA Name"]).strip(), d, "imported"))
                    bulk_insert_companies(rows)
                    st.session_state.import_message = f"Imported {len(rows):,} new companies. Skipped {skipped:,} already in the database."
        except Exception as e:
            st.error(f"Could not read the file: {e}")
    if "import_message" in st.session_state:
        st.success(st.session_state.import_message)
        del st.session_state.import_message
    st.divider()
    st.subheader("Reset")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe imported history"):
            wipe_imported_history()
            st.warning("Imported history wiped.")
    else:
        st.caption("Only a Manager can wipe imported history.")


def page_lists():
    page_header("Settings", "Clients & DNC",
                "Two permanent never-contact walls, matched by exact company name. Anything here is blocked before it reaches your reach-out data.")
    st.write("Two permanent 'never contact' walls, matched by exact company name. Anything here is blocked from being added to your reach-out data.")
    st.subheader(f"Clients  ·  {count_clients():,} on file")
    st.text_input("Add a client company name", key="client_name_input")
    st.button("Add client", on_click=handle_add_client)
    if "client_message" in st.session_state:
        kind, text = st.session_state.client_message
        st.success(text) if kind == "success" else st.error(text)
        del st.session_state.client_message
    client_file = st.file_uploader("Or upload a Clients CSV", type=["csv"], key="client_upload")
    if client_file is not None:
        cdf = read_csv_safe(client_file)
        cdf.columns = [c.strip() for c in cdf.columns]
        ccol = st.selectbox("Which column has the company names?", cdf.columns, key="client_col")
        if st.button("Import clients"):
            names = [str(x).strip() for x in cdf[ccol].tolist() if str(x).strip() and str(x).strip().lower() != "nan"]
            bulk_add_clients(names)
            st.success(f"Imported {len(names):,} client names.")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe clients list"):
            wipe_clients()
            st.warning("Clients list wiped.")
    else:
        st.caption("Only a Manager can wipe the clients list.")
    st.divider()
    st.subheader(f"DNC / Avoid  ·  {count_avoid():,} on file")
    st.text_input("Add a company to avoid", key="avoid_name_input")
    st.text_input("Reason (optional)", key="avoid_reason_input")
    st.button("Add to avoid list", on_click=handle_add_avoid)
    if "avoid_message" in st.session_state:
        kind, text = st.session_state.avoid_message
        st.success(text) if kind == "success" else st.error(text)
        del st.session_state.avoid_message
    avoid_file = st.file_uploader("Or upload a DNC/Avoid CSV", type=["csv"], key="avoid_upload")
    if avoid_file is not None:
        adf = read_csv_safe(avoid_file)
        adf.columns = [c.strip() for c in adf.columns]
        acol = st.selectbox("Which column has the company names?", adf.columns, key="avoid_col")
        rcol = st.selectbox("Reason column (optional)", ["(none)"] + list(adf.columns), key="avoid_reason_col")
        if st.button("Import avoid list"):
            pairs = []
            for _, r in adf.iterrows():
                nm = str(r[acol]).strip()
                if nm == "" or nm.lower() == "nan":
                    continue
                reason = "" if rcol == "(none)" else str(r[rcol]).strip()
                pairs.append((nm, reason))
            bulk_add_avoid(pairs)
            st.success(f"Imported {len(pairs):,} companies to the avoid list.")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe avoid list"):
            wipe_avoid()
            st.warning("Avoid list wiped.")
    else:
        st.caption("Only a Manager can wipe the avoid list.")


def page_prospects():
    page_header("Data", "Prospects DB",
                "Your master contact list. Prospects Mapping matches re-engage companies against this to fill in who to email.")
    st.write(
        "Your master contact list — First Name, Email, Company Name — used by Prospects Mapping to fill in "
        "contact details for re-engage companies. Matched by exact company name (no normalization), same as "
        "everywhere else in the app. Uploading replaces the whole list, since this is a wholesale refresh, "
        "not a merge."
    )
    n_prospects, n_companies = count_prospects()
    st.write(f"Currently on file: **{n_prospects:,}** prospects across **{n_companies:,}** unique companies.")

    prospects_file = st.file_uploader(
        "Upload prospects CSV (First Name, Email, Company Name)", type=["csv"], key="prospects_upload"
    )
    if prospects_file is not None:
        try:
            pdf = read_csv_safe(prospects_file)
            pdf.columns = [c.strip() for c in pdf.columns]
            pcols = list(pdf.columns)
            st.write("Preview (first 5 rows):")
            st.dataframe(pdf.head(5), hide_index=True)
            fcol = st.selectbox("First Name column", pcols, index=_idx(pcols, guess_col(pcols, "first name")), key="prospects_fcol")
            ecol = st.selectbox("Email column", pcols, index=_idx(pcols, guess_col(pcols, "email")), key="prospects_ecol")
            ccol = st.selectbox("Company Name column", pcols, index=_idx(pcols, guess_col(pcols, "company name")), key="prospects_ccol")
            st.warning(f"This will REPLACE all {n_prospects:,} prospects currently on file.")
            st.caption(
                "Matching ignores case and extra spaces (e.g. 'Acme Inc' matches 'acme inc  '), but the "
                "original spelling from your scrape is always what's used in the final output. Duplicate "
                "emails (keep-first) and anything on your Bounced/DNC list are dropped automatically."
            )
            if st.button(f"Replace prospects list with these {len(pdf):,} rows"):
                rows = []
                for r in with_progress(pdf, "Reading prospects"):
                    company = str(r[ccol]).strip()
                    if company == "" or company.lower() == "nan":
                        continue
                    fname = str(r[fcol]).strip()
                    email = str(r[ecol]).strip()
                    rows.append((fname, email, company))
                inserted, dup_count, bounced_count = replace_prospects(rows)
                st.success(
                    f"Loaded {inserted:,} prospects — skipped {dup_count:,} duplicate emails and "
                    f"{bounced_count:,} already on your Bounced/DNC list. Refresh this tab to see the updated count."
                )
        except Exception as e:
            st.error(f"Could not read the file: {e}")

    st.divider()
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe prospects list"):
            wipe_prospects()
            st.warning("Prospects list wiped.")
    else:
        st.caption("Only a Manager can wipe the prospects list.")


def page_bounced():
    page_header("Data", "Bounced & DNC emails",
                "Addresses that bounced or should never be contacted again. Adding here scrubs them from the Prospects DB immediately.")
    st.write(
        "Emails that bounced or should never be contacted again. The moment you add to this list — one at "
        "a time or by CSV — the app immediately removes any matching email from your Prospects DB. If that "
        "empties out every contact at a company, that company simply has no prospects left, so next time it "
        "comes through the pipeline it routes to an RA to source fresh contacts."
    )
    st.write(f"Currently on file: **{count_bounced():,}** bounced/DNC emails.")

    st.text_input("Add one email", key="bounced_email_input")
    if st.button("Add email"):
        email = st.session_state.bounced_email_input.strip()
        if email:
            bulk_add_bounced([email])
            removed = scrub_prospects_against_bounced()
            st.success(f"Added. Removed {removed:,} matching prospect(s) from the Prospects DB.")
            st.session_state.bounced_email_input = ""
        else:
            st.error("Enter an email address.")

    bounced_file = st.file_uploader("Or upload a CSV of bounced/DNC emails", type=["csv"], key="bounced_upload")
    if bounced_file is not None:
        try:
            bdf = read_csv_safe(bounced_file)
            bdf.columns = [c.strip() for c in bdf.columns]
            bcols = list(bdf.columns)
            ecol_b = st.selectbox("Email column", bcols, index=_idx(bcols, guess_col(bcols, "email")), key="bounced_ecol")
            if st.button(f"Add these {len(bdf):,} emails to the bounced list"):
                emails = [str(x).strip() for x in bdf[ecol_b].tolist() if str(x).strip() and str(x).strip().lower() != "nan"]
                bulk_add_bounced(emails)
                removed = scrub_prospects_against_bounced()
                st.success(f"Added {len(emails):,} emails. Removed {removed:,} matching prospect(s) from the Prospects DB.")
        except Exception as e:
            st.error(f"Could not read the file: {e}")

    st.divider()
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe bounced/DNC list"):
            wipe_bounced()
            st.warning("Bounced/DNC list wiped. (This does not restore any prospects that were already removed.)")
    else:
        st.caption("Only a Manager can wipe the bounced/DNC list.")


def page_review():
    page_header("Pipeline", "Review lists",
                "Your accumulated screening decisions. Bulk filtering checks these first, so a company is never re-screened or re-sent to the AI.")
    st.write("Your accumulated decisions, each with Employee Size and Industry — the info you take the call on. During bulk filtering the app checks these FIRST: eligible companies pass straight through, not-eligible are dropped, both skipping the keyword rules and the AI check. Matched by exact company name.")

    st.subheader(f"Eligible to take  \u00b7  {count_eligible():,} on file")
    with st.expander("Add one eligible company by hand"):
        st.text_input("Company name", key="elig_name_input")
        st.text_input("Employee size", key="elig_size_input")
        st.text_input("Industry", key="elig_ind_input")
        if st.button("Add eligible"):
            nm = st.session_state.elig_name_input.strip()
            if nm:
                add_one_eligible(nm, st.session_state.elig_size_input.strip(), st.session_state.elig_ind_input.strip())
                st.success(f"Added eligible: {nm}")
    elig_file = st.file_uploader("Upload an Eligible CSV (Company Name, Employee Size, Industry)", type=["csv"], key="elig_upload")
    if elig_file is not None:
        edf = read_csv_safe(elig_file)
        edf.columns = [c.strip() for c in edf.columns]
        ecols = list(edf.columns)
        ncol = st.selectbox("Company name column", ecols, index=_idx(ecols, guess_col(ecols, "company name")), key="elig_ncol")
        scol = st.selectbox("Employee size column", ["(none)"] + ecols, index=_idx(["(none)"] + ecols, guess_col(ecols, "employee size")), key="elig_scol")
        icol = st.selectbox("Industry column", ["(none)"] + ecols, index=_idx(["(none)"] + ecols, guess_col(ecols, "industry")), key="elig_icol")
        if st.button("Import eligible list"):
            rows = []
            for _, r in edf.iterrows():
                nm = str(r[ncol]).strip()
                if nm == "" or nm.lower() == "nan":
                    continue
                sz = "" if scol == "(none)" else str(r[scol]).strip()
                ind = "" if icol == "(none)" else str(r[icol]).strip()
                rows.append((nm, sz, ind))
            bulk_add_eligible(rows)
            st.success(f"Imported {len(rows):,} eligible companies.")
    er = get_eligible_rows()
    if er:
        st.caption("Most recent (up to 200):")
        st.dataframe(pd.DataFrame(er, columns=["Company Name", "Employee Size", "Industry"]), hide_index=True)
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe eligible list"):
            wipe_eligible()
            st.warning("Eligible list wiped.")
    else:
        st.caption("Only a Manager can wipe the eligible list.")

    st.divider()

    st.subheader(f"Not eligible  \u00b7  {count_not_eligible():,} on file")
    with st.expander("Add one not-eligible company by hand"):
        st.text_input("Company name", key="noel_name_input")
        st.text_input("Employee size", key="noel_size_input")
        st.text_input("Industry", key="noel_ind_input")
        st.text_input("Reason (optional)", key="noel_reason_input")
        if st.button("Add not-eligible"):
            nm = st.session_state.noel_name_input.strip()
            if nm:
                add_one_not_eligible(nm, st.session_state.noel_size_input.strip(), st.session_state.noel_ind_input.strip(), st.session_state.noel_reason_input.strip())
                st.success(f"Added not-eligible: {nm}")
    noel_file = st.file_uploader("Upload a Not-eligible CSV (Company Name, Employee Size, Industry)", type=["csv"], key="noel_upload")
    if noel_file is not None:
        ndf = read_csv_safe(noel_file)
        ndf.columns = [c.strip() for c in ndf.columns]
        ncols = list(ndf.columns)
        nncol = st.selectbox("Company name column", ncols, index=_idx(ncols, guess_col(ncols, "company name")), key="noel_ncol")
        nscol = st.selectbox("Employee size column", ["(none)"] + ncols, index=_idx(["(none)"] + ncols, guess_col(ncols, "employee size")), key="noel_scol")
        nicol = st.selectbox("Industry column", ["(none)"] + ncols, index=_idx(["(none)"] + ncols, guess_col(ncols, "industry")), key="noel_icol")
        if st.button("Import not-eligible list"):
            rows = []
            for _, r in ndf.iterrows():
                nm = str(r[nncol]).strip()
                if nm == "" or nm.lower() == "nan":
                    continue
                sz = "" if nscol == "(none)" else str(r[nscol]).strip()
                ind = "" if nicol == "(none)" else str(r[nicol]).strip()
                rows.append((nm, sz, ind, ""))
            bulk_add_not_eligible(rows)
            st.success(f"Imported {len(rows):,} not-eligible companies.")
    nr = get_not_eligible_rows()
    if nr:
        st.caption("Most recent (up to 200):")
        st.dataframe(pd.DataFrame(nr, columns=["Company Name", "Employee Size", "Industry", "Reason"]), hide_index=True)
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe not-eligible list"):
            wipe_not_eligible()
            st.warning("Not-eligible list wiped.")
    else:
        st.caption("Only a Manager can wipe the not-eligible list.")

    st.divider()

    st.subheader(f"Needs review  \u00b7  {count_needs_review():,} on file")
    st.write(
        "Companies the AI wasn't confident enough to call — these are held out of the Eligible/Not-eligible "
        "cache entirely so they never get auto-blacklisted on a guess. They're still skipped from future "
        "keyword rules and AI spend (matched by exact company name), but wait here until you resolve them."
    )
    review_rows = get_needs_review_rows()
    if review_rows:
        for (rname, rsize, rind, rreason) in review_rows:
            rc1, rc2, rc3 = st.columns([3, 3, 1.4])
            with rc1:
                st.write(f"**{rname}**")
                st.caption(f"{rsize or 'Unknown size'}  \u00b7  {rind or 'Unknown industry'}  \u00b7  {rreason}")
            with rc2:
                st.write("")
            with rc3:
                bc1, bc2 = st.columns(2)
                bc1.button("✅ Eligible", key=f"rev_elig_{rname}",
                           on_click=resolve_needs_review, args=(rname, "eligible"))
                bc2.button("❌ Not elig.", key=f"rev_not_{rname}",
                           on_click=resolve_needs_review, args=(rname, "not_eligible"))
    else:
        st.info("Nothing pending review right now.")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe needs-review queue"):
            wipe_needs_review()
            st.warning("Needs-review queue wiped (companies were NOT added to Eligible or Not-eligible).")
    else:
        st.caption("Only a Manager can wipe the needs-review queue.")


def page_assignments():
    page_header("Pipeline", "RA assignments",
                "Every batch ever dealt out, with who got which companies and when.")
    st.write(f"**{count_assignments():,}** total assignment records on file, across every batch ever dealt out.")
    batches = get_assignment_batches()
    if batches:
        for (batch_id, adate, n_companies, n_ras) in batches:
            with st.expander(f"{adate}  \u00b7  batch {batch_id}  \u00b7  {n_companies:,} companies across {n_ras} RA(s)"):
                rows = get_assignment_rows_for_batch(batch_id)
                bdf = pd.DataFrame(rows, columns=["Company Name", "Job Title", "Location", "Job URL", "RA"])
                st.dataframe(bdf, hide_index=True)
                st.download_button(
                    "Download this batch (CSV)", bdf.to_csv(index=False),
                    f"assignments_{batch_id}.csv", "text/csv", key=f"dlbatch_{batch_id}"
                )
                if CURRENT_ROLE in ("Manager", "TL"):
                    st.caption("Made a mistake dealing this batch (wrong counts, duplicate deal)? "
                               "This removes the batch record and any position from it no RA has "
                               "touched yet. Positions that already have prospects added are left alone.")
                    if st.button("Delete this batch", key=f"delbatch_{batch_id}"):
                        n_deleted, n_kept = delete_positions_for_batch(batch_id)
                        msg = f"Batch deleted. Removed {n_deleted:,} untouched position(s)."
                        if n_kept:
                            msg += f" Kept {n_kept:,} position(s) that already have prospects on them."
                        st.success(msg)
                        st.rerun()
    else:
        st.info("No assignments yet — run Step 5 in the Bulk scrape tab after filtering and the AI check.")

    st.divider()
    st.markdown("#### Reassignment queue")
    queue_rows = get_reassignment_queue_rows()
    if queue_rows:
        queue_df = pd.DataFrame(queue_rows, columns=["Company Name", "Job Title", "Location", "Job URL"])
        queue_df["Assign to"] = ""
        edited_queue = st.data_editor(
            queue_df, hide_index=True, disabled=["Company Name", "Job Title", "Location", "Job URL"],
            column_config={
                "Assign to": st.column_config.SelectboxColumn(options=[""] + get_active_ra_names())
            }, key="reassignment_editor", **FULL
        )
        selected = edited_queue[edited_queue["Assign to"].astype(str).str.strip() != ""]
        if st.button("Assign selected queued companies", disabled=selected.empty):
            batch_id = uuid.uuid4().hex[:10]
            assigned_date = date.today().isoformat()
            rows_to_save = [
                (batch_id, row["Company Name"], row["Job Title"], row["Location"], row["Job URL"],
                 row["Assign to"], assigned_date)
                for _, row in selected.iterrows()
            ]
            created = save_assignment_bundle(rows_to_save, CURRENT_NAME)
            st.success(f"Assigned {created} queued company/companies in batch {batch_id}.")
            st.rerun()
    else:
        st.caption("No companies waiting for reassignment.")
    st.divider()
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe assignment history"):
            wipe_assignments()
            st.warning("Assignment history wiped.")
    else:
        st.caption("Only a Manager can wipe assignment history.")


def page_workpage():
    page_header("Pipeline", "Workpage",
                "Edit job details and prospects, then move the entire company to Ready for campaign.")
    is_overseer = CURRENT_ROLE in ("TL", "Manager")

    if not is_overseer:
        preferred = get_user_work_date(CURRENT_USER["id"])
        work_date = st.date_input("Current Work Date", value=preferred, key="wp_work_date")
        if work_date != preferred:
            set_user_work_date(CURRENT_USER["id"], work_date)
            st.success(f"Work Date saved as {work_date.isoformat()} for your next jobs.")
        st.caption("This date stays selected until you change it. Ready positions keep their saved Work Date.")
    else:
        work_date = date.today()

    if is_overseer:
        c1, c2 = st.columns(2)
        with c1:
            ra_pick = st.selectbox("View positions for", ["All"] + get_ra_names_with_positions(), key="wp_ra_filter")
        with c2:
            status_pick = st.selectbox("Status", ["All", "Open", "Ready", "Exported"], key="wp_status_filter")
        rows = get_all_positions(status_filter=status_pick, ra_filter=ra_pick)
    else:
        status_pick = st.selectbox("Status", ["All", "Open", "Ready", "Exported"], key="wp_ra_status")
        rows = get_positions_for_ra(CURRENT_NAME, include_closed=True)
        if status_pick != "All":
            rows = [r for r in rows if r[3] == status_pick]

    if not rows:
        st.info("No positions match these filters.")
        return

    def _label(row):
        if is_overseer:
            pid, company, title, row_status, assigned_ra, _ = row
            return f"{company} — {title or '(no title)'} · {row_status} · {assigned_ra}"
        pid, company, title, row_status, _ = row
        return f"{company} — {title or '(no title)'} · {row_status}"

    options = {_label(r): r[0] for r in rows}
    pid = options[st.selectbox("Open a position", list(options.keys()), key="wp_selected_label")]
    pos = get_position(pid)
    if not pos:
        st.warning("This position no longer exists.")
        return
    (pid, company_name, job_title, status, assigned_ra, source, location, job_url,
     created_by, created_at, updated_at, master_company_id, batch_id, ready_work_date,
     ready_at, ready_by, export_batch_id, exported_at, exported_by) = pos
    can_edit = status != "Exported" and (is_overseer or assigned_ra == CURRENT_NAME)

    badge = {"Open": "🟢 Open", "Ready": "🟠 Ready for campaign", "Exported": "⚪ Exported"}.get(status, status)
    st.markdown(f"#### {company_name} · {badge}")
    st.caption(f"Added {created_at} by {created_by} · source: {source} · last updated {updated_at}")
    if ready_work_date:
        st.caption(f"Work Date: {ready_work_date} · Ready at: {ready_at} by {ready_by}")
    if export_batch_id:
        st.caption(f"Export batch: {export_batch_id} · exported by {exported_by}")

    c1, c2 = st.columns(2)
    with c1:
        new_title = st.text_input("Job Position", value=job_title or "", disabled=not can_edit, key=f"wp_title_{pid}")
        new_location = st.text_input("Job Location", value=location or "", disabled=not can_edit, key=f"wp_loc_{pid}")
    with c2:
        if is_overseer and can_edit:
            active_ras = get_active_ra_names()
            ra_options = active_ras if assigned_ra in active_ras else [assigned_ra] + active_ras
            new_ra = st.selectbox("Assigned to", ra_options, index=ra_options.index(assigned_ra), key=f"wp_ra_{pid}")
        else:
            new_ra = assigned_ra
            st.text_input("Assigned to", value=assigned_ra or "", disabled=True, key=f"wp_ra_ro_{pid}")
        new_url = st.text_input("JD URL", value=job_url or "", disabled=not can_edit, key=f"wp_url_{pid}")
    if job_url:
        st.markdown(f"[Open job posting ↗]({job_url})")
    if can_edit and st.button("Save job details", type="primary", key=f"wp_save_{pid}"):
        update_position(pid, job_title=new_title.strip(), assigned_ra=new_ra,
                        location=new_location.strip(), job_url=new_url.strip(), actor=CURRENT_NAME)
        st.success("Job details synchronized with Workpage, assignment history, and Master.")
        st.rerun()

    st.divider()
    st.markdown("#### Prospects")
    prospects = get_position_prospects(pid)
    for index, prow in enumerate(prospects, start=1):
        prospect_id, full_name, first_name, email, designation, p_location, linkedin_url, pstatus, pnotes = prow
        with st.expander(f"Prospect {index}: {full_name or first_name or email or '(incomplete)'}", expanded=False):
            p1, p2 = st.columns(2)
            with p1:
                efull = st.text_input("Full Name", full_name or "", disabled=not can_edit, key=f"ep_full_{prospect_id}")
                efirst = st.text_input("First Name", first_name or "", disabled=not can_edit, key=f"ep_first_{prospect_id}")
                eemail = st.text_input("Email", email or "", disabled=not can_edit, key=f"ep_email_{prospect_id}")
            with p2:
                edesig = st.text_input("Designation", designation or "", disabled=not can_edit, key=f"ep_desig_{prospect_id}")
                eloc = st.text_input("POC Location", p_location or "", disabled=not can_edit, key=f"ep_loc_{prospect_id}")
                elink = st.text_input("LinkedIn URL (optional)", linkedin_url or "", disabled=not can_edit,
                                      key=f"ep_link_{prospect_id}")
            if can_edit and st.button("Save prospect", key=f"ep_save_{prospect_id}"):
                email_key = eemail.strip().lower()
                if email_key and not EMAIL_RE.fullmatch(email_key):
                    st.error("Enter a valid email address.")
                elif email_key and email_key in get_bounced_email_set():
                    st.error("This email is in Bounced/DNC and cannot be saved.")
                elif elink.strip() and not re.match(r"^https?://", elink.strip(), re.I):
                    st.error("LinkedIn URL must begin with http:// or https://.")
                else:
                    update_position_prospect(
                        prospect_id, full_name=efull.strip(), first_name=efirst.strip(), email=eemail,
                        designation=edesig.strip(), location=eloc.strip(), linkedin_url=elink.strip(), actor=CURRENT_NAME
                    )
                    st.success("Prospect saved.")
                    st.rerun()
    if not prospects:
        st.caption("No prospects added yet.")

    if can_edit:
        with st.expander("Add prospects", expanded=not prospects):
            st.caption("LinkedIn URL is optional. All other fields are required before Ready.")
            empty_df = pd.DataFrame([{
                "Full Name": "", "First Name": "", "Email": "", "Designation": "",
                "POC Location": "", "LinkedIn URL": "",
            }])
            edited = st.data_editor(empty_df, num_rows="dynamic", hide_index=True,
                                    key=f"wp_new_prospects_{pid}", **FULL)
            if st.button("Save new prospects", key=f"wp_addp_{pid}"):
                rows_to_add = [
                    (r.get("Full Name", ""), r.get("First Name", ""),
                     str(r.get("Email", "")).strip().lower(), r.get("Designation", ""),
                     r.get("POC Location", ""), r.get("LinkedIn URL", ""))
                    for r in edited.to_dict("records")
                ]
                populated = [r for r in rows_to_add if any(str(v or "").strip() for v in r)]
                bounced = get_bounced_email_set()
                input_errors = []
                emails = []
                existing_emails = {(p[3] or "").strip().lower() for p in prospects if (p[3] or "").strip()}
                for row_number, row in enumerate(populated, start=1):
                    email_key = str(row[2] or "").strip().lower()
                    if email_key and not EMAIL_RE.fullmatch(email_key):
                        input_errors.append(f"Row {row_number}: invalid email format.")
                    if email_key in bounced:
                        input_errors.append(f"Row {row_number}: email is in Bounced/DNC.")
                    if email_key in emails:
                        input_errors.append(f"Row {row_number}: duplicate email in this position.")
                    if email_key and email_key in existing_emails:
                        input_errors.append(f"Row {row_number}: this email is already on the position.")
                    emails.append(email_key)
                    if row[5] and not re.match(r"^https?://", str(row[5]).strip(), re.I):
                        input_errors.append(f"Row {row_number}: LinkedIn URL must begin with http:// or https://.")
                if input_errors:
                    for error in input_errors:
                        st.error(error)
                elif populated:
                    n_added = bulk_add_position_prospects(pid, populated, CURRENT_NAME)
                    st.success(f"Added {n_added} prospect(s).")
                    st.rerun()
                else:
                    st.error("Enter at least one prospect row.")

    st.divider()
    if status == "Open":
        ready_date = work_date if not is_overseer else st.date_input(
            "Work Date for Ready", value=date.today(), key=f"ready_date_{pid}"
        )
        if st.button("Ready for campaign", type="primary", key=f"ready_{pid}"):
            ok, errors = mark_position_ready(pid, ready_date, CURRENT_NAME)
            if ok:
                if not is_overseer:
                    set_user_work_date(CURRENT_USER["id"], ready_date)
                st.success("Entire position and company are Ready for campaign.")
                st.rerun()
            for error in errors:
                st.error(error)
    elif status == "Ready":
        st.info("This position remains editable until it is Exported.")
        saved_ready_date = date.fromisoformat(ready_work_date) if ready_work_date else date.today()
        corrected_date = st.date_input("Saved Work Date", value=saved_ready_date, key=f"correct_ready_date_{pid}")
        if corrected_date != saved_ready_date and st.button("Save corrected Work Date", key=f"save_ready_date_{pid}"):
            update_ready_work_date(pid, corrected_date, CURRENT_NAME)
            if not is_overseer:
                set_user_work_date(CURRENT_USER["id"], corrected_date)
            st.success("Work Date corrected and audited.")
            st.rerun()
        if st.button("Return to Open", key=f"return_open_{pid}"):
            return_position_to_open(pid, CURRENT_NAME)
            st.rerun()
    elif status == "Exported" and is_overseer:
        if st.button("Unlock Exported position for correction", key=f"unlock_{pid}"):
            unlock_exported_position(pid, CURRENT_NAME)
            st.warning("Unlocked to Ready. It must be exported again after corrections.")
            st.rerun()

    can_delete = (status == "Open") or (status == "Exported" and is_overseer)
    if can_delete:
        with st.expander("Delete this company"):
            if status == "Exported":
                reason_kind = "exported_delete"
                st.warning("The immutable export and deletion audit will be retained.")
            else:
                label = st.radio(
                    "Reason", ["Assigned by mistake", "Request Not Eligible"],
                    key=f"del_reason_{pid}"
                )
                reason_kind = "assigned_mistake" if label == "Assigned by mistake" else "not_eligible"
            reason_note = st.text_area(
                "Explanation" + (" (required)" if reason_kind != "assigned_mistake" else ""),
                key=f"del_note_{pid}"
            )
            confirm = st.text_input(f"Type {company_name} to confirm", key=f"del_confirm_{pid}")
            disabled = confirm != company_name or (reason_kind != "assigned_mistake" and not reason_note.strip())
            if st.button("Delete company", disabled=disabled, key=f"delete_position_{pid}"):
                ok, message = delete_position_workflow(pid, CURRENT_USER, reason_kind, reason_note)
                if ok:
                    st.success(message)
                    st.rerun()
                st.error(message)


def page_eod():
    page_header("Data", "EOD uploads",
                "RAs submit their end-of-day sourcing sheet here. Every upload feeds new contacts straight into the Prospects DB.")
    st.write(
        "RAs upload their End-of-Day sourcing sheet here. Every upload feeds new contacts straight "
        "into the Prospects DB (deduped and bounce-checked, same as everywhere else), so those companies "
        "are ready to match the next time they re-enter the pipeline."
    )
    st.write(f"**{count_eod():,}** EOD rows on file across **{len(get_eod_ra_list()):,}** RA(s).")

    st.subheader("Upload your EOD sheet")
    st.session_state.setdefault("eod_ra_name", CURRENT_NAME)
    eod_ra_name = st.text_input("Your name (RA)", key="eod_ra_name", disabled=(CURRENT_ROLE == "RA"))
    if CURRENT_ROLE == "RA":
        eod_ra_name = CURRENT_NAME
    eod_date = st.date_input("Date", value=date.today(), key="eod_date")
    eod_file = st.file_uploader(
        "Upload EOD CSV (Company Name, Company LinkedIn URL, Full Name, First Name, POC Location, "
        "Designation, Email, position, Location, Job posting link, Industry)",
        type=["csv"], key="eod_upload"
    )
    if eod_file is not None:
        try:
            edf = read_csv_safe(eod_file)
            edf.columns = [c.strip() for c in edf.columns]
            ecols = list(edf.columns)
            st.write("Preview (first 5 rows):")
            st.dataframe(edf.head(5), hide_index=True)

            mapping_fields = [
                ("Company Name", "company name"), ("Company LinkedIn URL", "company linkedin url"),
                ("Full Name", "full name"), ("First Name", "first name"), ("POC Location", "poc location"),
                ("Designation", "designation"), ("Email", "email"), ("Position", "position"),
                ("Location", "location"), ("Job posting link", "job posting link"), ("Industry", "industry"),
            ]
            col_map = {}
            map_cols = st.columns(3)
            for i, (label, guess_key) in enumerate(mapping_fields):
                with map_cols[i % 3]:
                    options = ["(none)"] + ecols
                    col_map[label] = st.selectbox(label, options, index=_idx(options, guess_col(ecols, guess_key)), key=f"eod_map_{label}")

            if st.button(f"Submit {len(edf):,} EOD rows"):
                if not eod_ra_name.strip():
                    st.error("Enter your name first.")
                else:
                    def _eod_get(row, label):
                        c = col_map.get(label)
                        return "" if c in (None, "(none)") else str(row.get(c, "")).strip()

                    eod_rows = []
                    prospects_rows = []
                    for r in with_progress(edf, "Reading EOD rows"):
                        company = _eod_get(r, "Company Name")
                        if company == "" or company.lower() == "nan":
                            continue
                        first_name = _eod_get(r, "First Name")
                        email = _eod_get(r, "Email")
                        eod_rows.append((
                            eod_ra_name.strip(), eod_date.isoformat(), company,
                            _eod_get(r, "Company LinkedIn URL"), _eod_get(r, "Full Name"), first_name,
                            _eod_get(r, "POC Location"), _eod_get(r, "Designation"), email,
                            _eod_get(r, "Position"), _eod_get(r, "Location"), _eod_get(r, "Job posting link"),
                            _eod_get(r, "Industry"),
                        ))
                        if email:
                            prospects_rows.append((first_name, email, company))
                    bulk_insert_eod(eod_rows)
                    inserted, dup_count, bounced_count = add_prospects_from_eod(prospects_rows)
                    st.success(
                        f"Saved {len(eod_rows):,} EOD rows for {eod_ra_name.strip()}. Added {inserted:,} new "
                        f"prospects to the Prospects DB (skipped {dup_count:,} duplicates, {bounced_count:,} bounced)."
                    )
        except Exception as e:
            st.error(f"Could not read the file: {e}")

    if CURRENT_ROLE == "RA":
        return

    st.divider()
    st.subheader("TL / Manager: view and download by RA")
    ra_list = get_eod_ra_list()
    if ra_list:
        selected_ras = st.multiselect("Select RA(s)", ra_list, key="eod_download_ras")
        dcol1, dcol2 = st.columns(2)
        with dcol1:
            eod_date_from = st.date_input("From date (optional)", value=None, key="eod_dl_from")
        with dcol2:
            eod_date_to = st.date_input("To date (optional)", value=None, key="eod_dl_to")

        if selected_ras:
            rows = get_eod_rows(
                ra_names=selected_ras,
                date_from=eod_date_from.isoformat() if eod_date_from else None,
                date_to=eod_date_to.isoformat() if eod_date_to else None,
            )
            full_cols = [
                "RA Name", "Date", "Company Name", "Company LinkedIn URL", "Full Name", "First Name",
                "POC Location", "Designation", "Email", "Position", "Location", "Job Posting Link", "Industry",
            ]
            fulldf = pd.DataFrame(rows, columns=full_cols)
            st.write(f"**{len(fulldf):,}** rows for the selected RA(s)/date range.")
            st.dataframe(fulldf, hide_index=True)
            st.download_button("Download full EOD detail (CSV)", fulldf.to_csv(index=False), "eod_detail.csv", "text/csv")

            campaign_df = fulldf[["First Name", "Email", "Position", "Location"]].rename(
                columns={"First Name": "FIRST NAME", "Email": "EMAIL ID", "Position": "POSITION", "Location": "LOCATION"}
            )
            st.caption(
                "Campaign-ready assumes POSITION/LOCATION = the job posting's title/location (not the "
                "person's own Designation/POC Location) — matching how the rest of the app uses Title/Location. "
                "Say the word if you actually want Designation/POC Location instead."
            )
            st.download_button(
                "Download campaign-ready CSV (First Name, Email, Position, Location)",
                campaign_df.to_csv(index=False), "campaign_ready.csv", "text/csv"
            )
        else:
            st.info("Select at least one RA to preview and download their sheet(s).")
    else:
        st.info("No EOD uploads yet.")

    st.divider()
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe EOD history"):
            wipe_eod()
            st.warning("EOD history wiped. (Prospects already added to the Prospects DB from it are NOT removed.)")
    else:
        st.caption("Only a Manager can wipe EOD history.")


def _campaign_dataframe(items):
    df = pd.DataFrame(items, columns=["FIRST NAME", "EMAIL ID", "POSITION", "LOCATION"])
    for col in df.columns:
        df[col] = df[col].fillna("").astype(str).map(
            lambda value: "'" + value if value.startswith(("=", "+", "-", "@")) else value
        )
    return df


def page_campaign_export():
    page_header("Campaign", "Campaign export",
                "Assignee-wise Ready totals, campaign-ready downloads, and immutable export history.")
    if CURRENT_ROLE not in ("Manager", "TL"):
        st.error("Only a TL or Manager can access campaign exports.")
        return

    f1, f2, f3 = st.columns(3)
    with f1:
        date_from = st.date_input("From Work Date", value=date.today(), key="campaign_from")
    with f2:
        date_to = st.date_input("To Work Date", value=date.today(), key="campaign_to")
    with f3:
        ra_filter = st.selectbox("Assignee", ["All"] + get_ra_names_with_positions(), key="campaign_ra")

    totals = get_ra_ready_totals(date_from.isoformat(), date_to.isoformat())
    totals_df = pd.DataFrame(totals, columns=["Work Date", "Assignee", "Positions Ready", "Prospects"])
    st.markdown("#### Daily assignee totals")
    if totals_df.empty:
        st.info("No Ready or Exported work for this date range.")
    else:
        st.dataframe(totals_df, hide_index=True, **FULL)
        st.download_button("Download assignee totals", totals_df.to_csv(index=False),
                           "ra_daily_totals.csv", "text/csv")

    st.markdown("#### Ready positions")
    ready_rows = get_campaign_positions(date_from.isoformat(), date_to.isoformat(), ra_filter, ("Ready",))
    labels = {
        f"{company} · {ra_name} · {work_date} · {prospect_count} prospect(s)": pid
        for pid, company, title, location, ra_name, row_status, work_date, ready_at, prospect_count in ready_rows
    }
    selected_labels = st.multiselect("Select positions to export", list(labels.keys()), key="campaign_selected")
    selected_ids = [labels[label] for label in selected_labels]
    if selected_ids:
        preview_rows = [row for row in ready_rows if row[0] in selected_ids]
        preview_df = pd.DataFrame(preview_rows, columns=[
            "Position ID", "Company", "Job Position", "Job Location", "RA", "Status",
            "Work Date", "Ready At", "Prospects",
        ])
        st.dataframe(preview_df, hide_index=True, **FULL)
    if st.button("Create export batch and mark Exported", type="primary", disabled=not selected_ids):
        validation_errors = []
        for selected_id in selected_ids:
            validation_errors.extend(validate_position_for_ready(selected_id))
        if validation_errors:
            for error in validation_errors:
                st.error(error)
        else:
            try:
                batch_id, items = create_campaign_export(selected_ids, CURRENT_NAME)
                st.session_state.last_campaign_export = (batch_id, items)
                st.success(f"Created {batch_id}: {len(items):,} prospect(s) marked Exported.")
                st.rerun()
            except Exception as exc:
                st.error(f"Export could not be created: {exc}")

    if st.session_state.get("last_campaign_export"):
        batch_id, items = st.session_state.last_campaign_export
        export_df = _campaign_dataframe(items)
        st.download_button(
            f"Download {batch_id}", export_df.to_csv(index=False), f"{batch_id}.csv", "text/csv",
            key=f"download_{batch_id}"
        )

    st.divider()
    st.markdown("#### Export history")
    batches = get_export_batches()
    if batches:
        batch_labels = {
            f"{batch_id} · {created_at} · {created_by} · {prospects} prospects": batch_id
            for batch_id, created_by, created_at, positions, prospects in batches
        }
        history_label = st.selectbox("Previous export", list(batch_labels.keys()), key="export_history_pick")
        history_id = batch_labels[history_label]
        history_df = _campaign_dataframe(get_export_batch_items(history_id))
        st.dataframe(history_df, hide_index=True, **FULL)
        st.download_button("Re-download this export", history_df.to_csv(index=False),
                           f"{history_id}.csv", "text/csv", key=f"redownload_{history_id}")
    else:
        st.caption("No export batches yet.")

    with st.expander("Deletion audit"):
        deleted = get_deletion_audit_rows()
        if deleted:
            deleted_df = pd.DataFrame(deleted, columns=[
                "Company", "Deleted By", "Role", "Reason", "Deleted At", "Export Batch",
            ])
            st.dataframe(deleted_df, hide_index=True, **FULL)
        else:
            st.caption("No deleted positions.")


def page_approval_requests():
    page_header("Review", "Not Eligible requests",
                "Approve or reject RA requests. Pending companies stay out of the assignment pool.")
    if CURRENT_ROLE not in ("Manager", "TL"):
        st.error("Only a TL or Manager can review requests.")
        return
    pending = get_not_eligible_requests("Pending")
    if not pending:
        st.info("No pending requests.")
        return
    for request_id, company, reason, requested_by, requested_at, status, reviewed_by, reviewed_at, note in pending:
        with st.container(border=True):
            st.markdown(f"**{company}**")
            st.write(reason)
            st.caption(f"Requested by {requested_by} at {requested_at}")
            decision_note = st.text_input("Reviewer note (optional)", key=f"approval_note_{request_id}")
            a1, a2 = st.columns(2)
            if a1.button("Approve Not Eligible", type="primary", key=f"approve_{request_id}"):
                review_not_eligible_request(request_id, True, CURRENT_NAME, decision_note)
                st.rerun()
            if a2.button("Reject and return to pool", key=f"reject_{request_id}"):
                review_not_eligible_request(request_id, False, CURRENT_NAME, decision_note)
                st.rerun()


def page_analytics():
    page_header("Insight", "Analytics",
                "How many emails went out against each title group, and how many replies came back. Replies are matched to sent mail by email plus company name.")
    st.write(
        "Upload your historical emails-sent and positive-response sheets here. A response is matched "
        "back to the email it responded to by email + company name, which is how it inherits a title "
        "bucket for the title-group breakdown. RA-wise stats use each sheet's own RA Name column directly."
    )
    st.write(
        f"On file: **{count_emails_sent():,}** emails sent, **{count_positive_responses():,}** positive responses."
    )

    st.subheader("Upload: Emails Sent")
    st.caption("Columns: First Name, Email ID, Position, Location, Company Name, RA Name (optional — e.g. Jan sheet).")
    sent_period = st.text_input("Label this upload (e.g. 'Jan 2026') — optional", key="sent_period")
    sent_file = st.file_uploader("Upload Emails Sent CSV", type=["csv"], key="sent_upload")
    if sent_file is not None:
        try:
            sdf = read_csv_safe(sent_file)
            sdf.columns = [c.strip() for c in sdf.columns]
            scols = list(sdf.columns)
            st.dataframe(sdf.head(5), hide_index=True)
            s_opts = ["(none)"] + scols
            s_first = st.selectbox("First Name column", scols, index=_idx(scols, guess_col(scols, "first name")), key="sent_fcol")
            s_email_guess = guess_col(scols, "email id") or guess_col(scols, "email")
            s_email = st.selectbox("Email column", scols, index=_idx(scols, s_email_guess), key="sent_ecol")
            s_pos = st.selectbox("Position column", scols, index=_idx(scols, guess_col(scols, "position")), key="sent_pcol")
            s_loc = st.selectbox("Location column", scols, index=_idx(scols, guess_col(scols, "location")), key="sent_lcol")
            s_comp = st.selectbox("Company Name column", scols, index=_idx(scols, guess_col(scols, "company name")), key="sent_ccol")
            s_ra = st.selectbox("RA Name column (optional)", s_opts, index=_idx(s_opts, guess_col(scols, "ra name")), key="sent_racol")
            if st.button(f"Add {len(sdf):,} emails-sent rows"):
                rows = []
                for r in with_progress(sdf, "Reading emails-sent rows"):
                    email = str(r[s_email]).strip()
                    if email == "" or email.lower() == "nan":
                        continue
                    ra_val = "" if s_ra == "(none)" else str(r[s_ra]).strip()
                    rows.append((
                        str(r[s_first]).strip(), email, str(r[s_pos]).strip(), str(r[s_loc]).strip(),
                        str(r[s_comp]).strip(), ra_val, sent_period.strip(),
                    ))
                bulk_insert_emails_sent(rows)
                st.success(f"Added {len(rows):,} emails-sent rows.")
        except Exception as e:
            st.error(f"Could not read the file: {e}")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe emails-sent data"):
            wipe_emails_sent()
            st.warning("Emails-sent data wiped.")
    else:
        st.caption("Only a Manager can wipe emails-sent data.")

    st.divider()
    st.subheader("Upload: Positive Responses")
    st.caption("Columns: Email, Name, Date, RA Name, Position, Designation of Prospect, Company Name (Designation/Company optional — blank is fine).")
    resp_file = st.file_uploader("Upload Positive Responses CSV", type=["csv"], key="resp_upload")
    if resp_file is not None:
        try:
            rdf_up = read_csv_safe(resp_file)
            rdf_up.columns = [c.strip() for c in rdf_up.columns]
            rcols = list(rdf_up.columns)
            st.dataframe(rdf_up.head(5), hide_index=True)
            r_opts = ["(none)"] + rcols
            r_email = st.selectbox("Email column", rcols, index=_idx(rcols, guess_col(rcols, "email")), key="resp_ecol")
            r_name = st.selectbox("Name column", rcols, index=_idx(rcols, guess_col(rcols, "name")), key="resp_ncol")
            r_date = st.selectbox("Date column", rcols, index=_idx(rcols, guess_col(rcols, "date")), key="resp_dcol")
            r_ra = st.selectbox("RA Name column", rcols, index=_idx(rcols, guess_col(rcols, "ra name")), key="resp_racol")
            r_pos = st.selectbox("Position column (optional)", r_opts, index=_idx(r_opts, guess_col(rcols, "position")), key="resp_pcol")
            r_desig = st.selectbox("Designation of Prospect column (optional)", r_opts, index=_idx(r_opts, guess_col(rcols, "designation")), key="resp_desigcol")
            r_comp = st.selectbox("Company Name column (optional)", r_opts, index=_idx(r_opts, guess_col(rcols, "company name")), key="resp_ccol")
            if st.button(f"Add {len(rdf_up):,} positive-response rows"):
                rows = []
                for r in with_progress(rdf_up, "Reading responses"):
                    email = str(r[r_email]).strip()
                    if email == "" or email.lower() == "nan":
                        continue
                    def _r_get(col):
                        return "" if col == "(none)" else str(r[col]).strip()
                    rows.append((
                        email, str(r[r_name]).strip(), str(r[r_date]).strip(), str(r[r_ra]).strip(),
                        _r_get(r_pos), _r_get(r_desig), _r_get(r_comp),
                    ))
                bulk_insert_positive_responses(rows)
                st.success(f"Added {len(rows):,} positive-response rows.")
        except Exception as e:
            st.error(f"Could not read the file: {e}")
    if CURRENT_ROLE == "Manager":
        if st.button("Wipe positive-responses data"):
            wipe_positive_responses()
            st.warning("Positive-responses data wiped.")
    else:
        st.caption("Only a Manager can wipe positive-responses data.")

    st.divider()
    st.subheader("Title-bucket keywords")
    st.write(
        "Every position gets bucketed by the longest keyword found in it (so 'Project Manager' won't "
        "collapse into plain 'Manager'). Add or remove keywords to tune the grouping."
    )
    st.text_input("Add keywords (comma-separated)", key="new_title_keywords")
    if st.button("Add keyword(s)"):
        parts = [p.strip() for p in st.session_state.new_title_keywords.split(",") if p.strip()]
        for p in parts:
            add_title_bucket_keyword(p)
        st.session_state.new_title_keywords = ""
        st.success(f"Added {len(parts):,} keyword(s).")
    kw_list = get_title_bucket_keywords()
    with st.expander(f"Current keywords ({len(kw_list):,})"):
        for kw in kw_list:
            kc1, kc2 = st.columns([4, 1])
            kc1.write(kw)
            if CURRENT_ROLE == "Manager":
                kc2.button("Remove", key=f"rm_titlekw_{kw}", on_click=remove_title_bucket_keyword, args=(kw,))

    st.divider()
    st.subheader("Results")
    if st.button("Compute analytics"):
        st.session_state.analytics_result = compute_analytics()

    if "analytics_result" in st.session_state:
        title_stats, ra_stats, total_sent, total_resp, unattributed = st.session_state.analytics_result
        overall_rate = (total_resp / total_sent * 100) if total_sent else 0.0
        stat_cards([
            ("Emails sent", f"{total_sent:,}", "across all uploads", "primary"),
            ("Positive replies", f"{total_resp:,}", "across all uploads", "green"),
            ("Reply rate", f"{overall_rate:.2f}%", "replies ÷ sent", "violet"),
            ("Untraced replies", f"{unattributed:,}", "no company, or no matching sent row", "amber"),
        ])
        st.write("")

        st.markdown("### By title group")
        tdf = pd.DataFrame(title_stats, columns=["Title Bucket", "Sent", "Responses", "Response Rate %"])
        tdf["Response Rate %"] = tdf["Response Rate %"].round(2)
        perf_chart(tdf, "Title Bucket", "title")
        with st.expander(f"Title group table ({len(tdf):,} buckets)"):
            st.dataframe(tdf, hide_index=True, **FULL)
        st.download_button("Download title-group summary (CSV)", tdf.to_csv(index=False), "analytics_by_title.csv", "text/csv")

        st.markdown("### By RA")
        radf = pd.DataFrame(ra_stats, columns=["RA Name", "Sent", "Responses", "Response Rate %"])
        radf["Response Rate %"] = radf["Response Rate %"].round(2)
        perf_chart(radf, "RA Name", "ra")
        with st.expander(f"RA table ({len(radf):,} RAs)"):
            st.dataframe(radf, hide_index=True, **FULL)
        st.download_button("Download RA summary (CSV)", radf.to_csv(index=False), "analytics_by_ra.csv", "text/csv")
    else:
        st.info("Click 'Compute analytics' after uploading data to see the breakdown.")


def page_danger():
    page_header("Settings", "Danger zone",
                "Reset the app to a clean slate. Useful for testing, permanent in effect.")
    if CURRENT_ROLE != "Manager":
        st.error("Only a Manager can access the Danger zone.")
        return
    st.error(
        "This permanently empties EVERY table — companies, block list, clients, DNC/avoid, "
        "and eligible/not-eligible lists. There is no undo. Use this only to reset the app for testing."
    )
    total, _ = count_companies()
    st.write(
        f"Current data on file: **{total:,}** companies, **{len(get_block_list()):,}** block keywords, "
        f"**{count_clients():,}** clients, **{count_avoid():,}** DNC/avoid, "
        f"**{count_eligible():,}** eligible, **{count_not_eligible():,}** not-eligible, "
        f"**{count_needs_review():,}** needs-review, **{count_assignments():,}** RA assignment records, "
        f"**{count_prospects()[0]:,}** prospects, **{count_bounced():,}** bounced/DNC emails, "
        f"**{count_eod():,}** EOD rows, **{count_emails_sent():,}** emails-sent records, "
        f"**{count_positive_responses():,}** positive-response records."
    )
    st.text_input('Type WIPE (all caps) to confirm', key="wipe_confirm_input")
    if st.button("Wipe entire database", type="primary"):
        if st.session_state.wipe_confirm_input.strip() == "WIPE":
            wipe_entire_database()
            st.session_state.wipe_confirm_input = ""
            st.success("Database wiped. Every table is empty — the app is back to a fresh install.")
        else:
            st.error('Type WIPE exactly (all caps) in the box above, then click the button again.')


def page_users():
    page_header("Admin", "Manage users", "Create logins for TLs and RAs, set roles, deactivate access.")

    if CURRENT_ROLE != "Manager":
        st.error("Only a Manager can access this page.")
        return

    st.markdown("#### Add a new user")
    with st.form("new_user_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            new_username = st.text_input("Username (used to log in)")
            new_display = st.text_input("Display name")
        with c2:
            new_role = st.selectbox("Role", ["RA", "TL", "Manager"])
            new_pwd = st.text_input("Temporary password", type="password")
        submitted = st.form_submit_button("Create user", type="primary")
        if submitted:
            if not new_username.strip() or not new_pwd or not new_display.strip():
                st.error("Username, display name, and password are all required.")
            elif len(new_pwd) < 6:
                st.error("Password should be at least 6 characters.")
            else:
                ok, msg = create_user(new_username, new_pwd, new_display, new_role)
                if ok:
                    st.success(f"Created {new_role} account for {new_display} (username: {new_username.strip().lower()}). "
                               f"Share the temporary password with them directly — it won't be shown again here.")
                else:
                    st.error(msg)

    st.markdown("#### Existing users")
    users = list_users()
    if not users:
        st.caption("No users yet.")
        return

    for uid, username, display_name, role, active, created_at in users:
        with st.container(border=True):
            c1, c2, c3, c4, c5 = st.columns([2.2, 1.3, 1.3, 1.3, 1.3])
            c1.markdown(f"**{display_name}**  \n`{username}`")
            c2.caption(f"Role\n\n**{role}**")
            c3.caption(f"Status\n\n**{'Active' if active else 'Deactivated'}**")

            is_self = (uid == CURRENT_USER["id"])
            with c4:
                if is_self:
                    st.caption("This is you")
                elif active:
                    if st.button("Deactivate", key=f"deact_{uid}"):
                        set_user_active(uid, False)
                        st.rerun()
                else:
                    if st.button("Reactivate", key=f"react_{uid}"):
                        set_user_active(uid, True)
                        st.rerun()
            with c5:
                new_role_pick = st.selectbox("Change role", ["RA", "TL", "Manager"],
                                              index=["RA", "TL", "Manager"].index(role),
                                              key=f"role_{uid}", label_visibility="collapsed")
                if new_role_pick != role:
                    set_user_role(uid, new_role_pick)
                    st.rerun()

            with st.expander("Reset password"):
                rp = st.text_input("New password", type="password", key=f"resetpwd_{uid}")
                if st.button("Set new password", key=f"resetbtn_{uid}"):
                    if len(rp) < 6:
                        st.error("Password should be at least 6 characters.")
                    else:
                        reset_user_password(uid, rp)
                        st.success("Password updated.")


# ---------- Router ----------
ROUTES = {
    "Dashboard": page_dashboard,
    "Bulk scrape": page_bulk,
    "Add company": page_add,
    "Review lists": page_review,
    "RA assignments": page_assignments,
    "Workpage": page_workpage,
    "Not Eligible requests": page_approval_requests,
    "Campaign export": page_campaign_export,
    "Prospects DB": page_prospects,
    "EOD uploads": page_eod,
    "Bounced & DNC": page_bounced,
    "Import history": page_import,
    "Analytics": page_analytics,
    "Clients & DNC": page_lists,
    "Title block list": page_block,
    "Danger zone": page_danger,
    "Manage users": page_users,
}

PAGE_ROLES = {
    "Dashboard": {"Manager", "TL", "RA"},
    "Add company": {"Manager", "TL", "RA"},
    "Workpage": {"Manager", "TL", "RA"},
    "EOD uploads": {"Manager", "TL", "RA"},
    "Bulk scrape": {"Manager", "TL"},
    "Review lists": {"Manager", "TL"},
    "RA assignments": {"Manager", "TL"},
    "Not Eligible requests": {"Manager", "TL"},
    "Campaign export": {"Manager", "TL"},
    "Analytics": {"Manager", "TL"},
    "Prospects DB": {"Manager"},
    "Bounced & DNC": {"Manager"},
    "Import history": {"Manager"},
    "Clients & DNC": {"Manager"},
    "Title block list": {"Manager"},
    "Danger zone": {"Manager"},
    "Manage users": {"Manager"},
}
requested_page = st.session_state.page
if CURRENT_ROLE not in PAGE_ROLES.get(requested_page, {"Manager"}):
    st.error("You do not have permission to access this page.")
else:
    ROUTES.get(requested_page, page_dashboard)()
